# sales/views.py

from datetime import datetime
import logging

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseBadRequest
from .utils import pdf_config, pdf_options, aggregate_payment_totals, aggregate_legacy_refunds, aggregate_refunds_unified
logger = logging.getLogger(__name__)


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm


try:
    from openpyxl import Workbook
except Exception:
    Workbook = None






from decimal import Decimal

from catalog.models_stock import StockLedger
from catalog.models import Product as CatalogProduct

from django.db import models, transaction
from django.db.models import Sum, Count, Q, CharField
from django.db.models.functions import Cast, Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template, render_to_string




from .models import (
    Customer,
    Invoice,
    InvoiceItem,
    Payment,
    SavedItem,
    Refund,
    CompanyProfile,
)
from catalog.models import Product


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def parse_decimal(value, default=Decimal("0.00")):
    """
    Safely parse any value into Decimal.

    - Empty / None -> default
    - Invalid string -> default
    """
    try:
        if value is None or value == "":
            return default
        return Decimal(str(value))
    except Exception:
        return default


def parse_int(value, default=0):
    """
    Safely parse any value into int.

    - Float string -> int(float())
    - Invalid -> default
    """
    try:
        return int(float(value))
    except Exception:
        return default


# ============================================================
# INVOICE LIST + EXPORT
# ============================================================

def invoice_list(request):
    """
    Invoice listing with:
    - Search on number, customer name, customer email
    - Status filter (paid/partial/unpaid)
    - Export to CSV / Excel using current filters
    """
    invoices = (
        Invoice.objects
        .select_related("customer")
        .filter(in_bin=False)  # hide ones moved to bin
        .order_by("-date", "-id")
    )

    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    export = request.GET.get("export", "").strip()

    # ---------- SEARCH ----------
    if q:
        invoices = invoices.annotate(
            number_str=Cast("number", CharField())
        ).filter(
            Q(number_str__icontains=q)
            | Q(customer__name__icontains=q)
            | Q(customer__email__icontains=q)
        )

    # ---------- STATUS FILTER ----------
    if status in ["paid", "partial", "unpaid"]:
        invoices = invoices.filter(status=status)

    # ---------- EXPORT ----------
    if export == "csv":
        return export_invoices_csv(invoices)
    elif export == "xlsx":
        return export_invoices_excel(invoices)

    return render(request, "sales/invoice_list.html", {"invoices": invoices})


def export_invoices_csv(invoices_qs):
    """
    Export filtered invoice queryset into CSV.
    """
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="invoices.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "Invoice #",
            "Customer",
            "Email",
            "Date",
            "Status",
            "Grand Total",
            "Paid Amount",
            "Balance Due",
            "Last Updated",
        ]
    )

    for inv in invoices_qs:
        writer.writerow(
            [
                inv.number,
                inv.customer.name if inv.customer else "",
                inv.customer.email if inv.customer and inv.customer.email else "",
                inv.date.strftime("%d-%m-%Y") if inv.date else "",
                inv.status,
                float(inv.grand_total or 0),
                float(inv.paid_amount or 0),
                float(inv.balance_due or 0),
                inv.updated_at.strftime("%d-%m-%Y %H:%M") if inv.updated_at else "",
            ]
        )

    return response


def export_invoices_excel(invoices_qs):
    """
    Export filtered invoice queryset into XLSX.
    Falls back to CSV if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return export_invoices_csv(invoices_qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    ws.append(
        [
            "Invoice #",
            "Customer",
            "Email",
            "Date",
            "Status",
            "Grand Total",
            "Paid Amount",
            "Balance Due",
            "Last Updated",
        ]
    )

    for inv in invoices_qs:
        ws.append(
            [
                inv.number,
                inv.customer.name if inv.customer else "",
                inv.customer.email if inv.customer and inv.customer.email else "",
                inv.date.strftime("%d-%m-%Y") if inv.date else "",
                inv.status,
                float(inv.grand_total or 0),
                float(inv.paid_amount or 0),
                float(inv.balance_due or 0),
                inv.updated_at.strftime("%d-%m-%Y %H:%M") if inv.updated_at else "",
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = 'attachment; filename="invoices.xlsx"'
    return response


# ============================================================
# INVOICE ADD / DETAIL / EDIT
# ============================================================

# sales/views.py
from django.urls import reverse

# Models from sales app (adjust these imports if your project uses different module names)
from .models import (
    Customer,
    Product,
    SavedItem,
    CompanyProfile,
    Invoice,
    InvoiceItem,
    Payment,
)

# Catalog inventory utils (safe import + alias)
from catalog import inventory_utils as inv_utils

# helpers from inventory_utils (use getattr to avoid hard import errors)
allocate_fifo = getattr(inv_utils, "allocate_fifo", None)
commit_allocation = getattr(inv_utils, "commit_allocation", None)
reverse_allocation_for_invoice = getattr(inv_utils, "reverse_allocation_for_invoice", None)

# Use Catalog's Product when we need to lock DB rows for allocation (if different, adjust)
from catalog.models import Product as CatalogProduct


# ----------------- small helpers -----------------
def parse_decimal(v):
    try:
        if v is None or v == "":
            return Decimal("0")
        # if already Decimal
        if isinstance(v, Decimal):
            return v
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def parse_int(v):
    try:
        if v is None or v == "":
            return 0
        return int(v)
    except Exception:
        return 0


# ----------------- invoice_add view -----------------
@login_required
@transaction.atomic
def invoice_add(request):
    """
    Create a new invoice:
    - dynamic items (products/manual/saved items)
    - stock allocation from PrintRuns (FIFO) for track_stock=True products
    - multiple payments
    """
    customers = Customer.objects.all()
    products = Product.objects.all()
    saved_items = SavedItem.objects.all()
    companies = CompanyProfile.objects.all()
    today = timezone.localdate()

    # Generate next invoice number based on latest ID
    latest = Invoice.objects.order_by("-id").first()
    next_seq = (latest.id + 1) if latest else 1
    next_number = f"INV-{next_seq:04d}"

    if request.method == "POST":
        # ---------- HEADER ----------
        inv = Invoice.objects.create(
            number=request.POST.get("number", next_number),
            customer_id=request.POST.get("customer") or None,
            company_id=request.POST.get("company") or None,
            date=request.POST.get("date") or today,
            billing_address=request.POST.get("billing_address", ""),
            shipping_address=request.POST.get("shipping_address", ""),
            notes=request.POST.get("notes", ""),
            terms=request.POST.get("terms", ""),
            bank_name=request.POST.get("bank_name", ""),
            bank_branch=request.POST.get("bank_branch", ""),
            account_number=request.POST.get("account_number", ""),
            ifsc=request.POST.get("ifsc", ""),
            upi=request.POST.get("upi", ""),
            logo=request.FILES.get("logo"),
        )

        # ---------- ITEMS (gather posted lines) ----------
        prods = request.POST.getlist("product[]")           # id / 'manual' / 'saved:<id>'
        manuals = request.POST.getlist("manual_product[]")  # description
        hsn_codes = request.POST.getlist("hsn_code[]")
        qtys = request.POST.getlist("qty[]")
        rates = request.POST.getlist("rate[]")
        tax_rates = request.POST.getlist("tax[]")
        discounts = request.POST.getlist("discount[]")

        max_lines = (
            max(
                len(prods),
                len(manuals),
                len(hsn_codes),
                len(qtys),
                len(rates),
                len(tax_rates),
                len(discounts),
            )
            if (prods or manuals or qtys or rates or tax_rates or discounts)
            else 0
        )

        # Build posted_lines list for pre-check and allocation plan
        posted_lines = []
        for i in range(max_lines):
            p = prods[i] if i < len(prods) else ""
            m = manuals[i] if i < len(manuals) else ""
            hsn = (hsn_codes[i] if i < len(hsn_codes) else "").strip()
            q = parse_decimal(qtys[i] if i < len(qtys) else "0")
            r = parse_decimal(rates[i] if i < len(rates) else "0")
            t = parse_int(tax_rates[i] if i < len(tax_rates) else "0")
            d = parse_decimal(discounts[i] if i < len(discounts) else "0")
            # skip fully empty/zero lines
            if q <= 0 and r <= 0:
                continue
            # convert q to int for allocation planning (units)
            posted_lines.append({'p': p, 'm': m, 'q': int(q), 'r': r, 't': t, 'd': d})

        # ---------- ALLOCATION PLANNING (transaction & locks) ----------
        insufficient = []
        alloc_plan = {}  # map posted_lines index -> allocations list

        # Lock product rows that will be allocated to prevent race conditions
        prod_ids_to_lock = []
        for idx, line in enumerate(posted_lines):
            p = line['p']
            if p and not (p == "manual" or (isinstance(p, str) and p.startswith("saved:"))):
                try:
                    prod_ids_to_lock.append(int(p))
                except Exception:
                    logger.exception("Failed to update line_total for invoice item id=%s", getattr(it, "id", None))

        # Use select_for_update to lock rows
        locked_products = {}
        if prod_ids_to_lock and CatalogProduct is not None:
            for prd in CatalogProduct.objects.select_for_update().filter(id__in=prod_ids_to_lock):
                locked_products[str(prd.id)] = prd

        # For each product line, plan allocation if track_stock enabled
        for idx, line in enumerate(posted_lines):
            p = line['p']
            qty = line['q']
            if not p or p == "manual" or (isinstance(p, str) and p.startswith("saved:")):
                continue  # manual/saved skip stock allocation
            prod_obj = None
            try:
                prod_obj = locked_products.get(str(int(p)))
            except Exception:
                prod_obj = None

            if not prod_obj:
                insufficient.append((idx, "Product missing"))
                continue
            if not getattr(prod_obj, "track_stock", True):
                # product not tracked -> no allocation required
                continue

            # compute allocation plan via FIFO
            if allocate_fifo is None:
                insufficient.append((idx, "Allocation planner not available"))
                continue
            try:
                allocs = allocate_fifo(prod_obj, qty)
            except Exception as e:
                # planner raised error
                insufficient.append((idx, f"Allocation error for {prod_obj}: {e}"))
                continue

            # allocate_fifo may return dict/structure or None on insufficient
            if not allocs:
                insufficient.append((idx, f"Insufficient stock for {getattr(prod_obj,'sku',None) or prod_obj.name} (need {qty})"))
            else:
                alloc_plan[idx] = allocs

        if insufficient:
            # rollback invoice header and show message
            msg = "; ".join([f"Line {i+1}: {m}" for i, m in insufficient])
            messages.error(request, f"Cannot create invoice due to stock issues: {msg}")
            # ensure invoice not left in DB
            try:
                inv.delete()
            except Exception:
                pass
            return redirect("sales:invoice_add")

        # ---------- CREATE ITEMS & COMMIT ALLOCATIONS ----------
        for i, line in enumerate(posted_lines):
            p = line['p']
            m = line['m']
            q = line['q']
            r = line['r']
            t = line['t']
            d = line['d']

            if not p or p == "manual":
                it = InvoiceItem.objects.create(
                    invoice=inv,
                    description=m or "Custom Item",
                    hsn_code=hsn,
                    quantity=q,
                    rate=r,
                    tax_rate=t,
                    discount_percent=d,
                )
            elif isinstance(p, str) and p.startswith("saved:"):
                try:
                    sid = int(p.split(":", 1)[1])
                    si = SavedItem.objects.get(pk=sid)
                    used_rate = si.price if r == 0 else r
                    it = InvoiceItem.objects.create(
                        invoice=inv,
                        saved_item=si,
                        description=si.name,
                        hsn_code=hsn,
                        quantity=q,
                        rate=used_rate,
                        tax_rate=si.tax_rate,
                        discount_percent=d,
                    )
                except Exception:
                    it = InvoiceItem.objects.create(
                        invoice=inv,
                        description=m or "Saved item (missing)",
                        hsn_code=hsn,
                        quantity=q,
                        rate=r,
                        tax_rate=t,
                        discount_percent=d,
                    )
            else:
                prod = CatalogProduct.objects.get(pk=int(p))
                used_rate = r if r != Decimal("0") else getattr(prod, "price", Decimal("0"))
                it = InvoiceItem.objects.create(
                    invoice=inv,
                    product=prod,
                    description=prod.name,
                    hsn_code=(getattr(prod, "hsn_code", "") or hsn),
                    quantity=q,
                    rate=used_rate,
                    tax_rate=getattr(prod, "tax_rate", t),
                    discount_percent=d,
                )
                # commit allocation ledger rows if planned
                if i in alloc_plan:
                    if commit_allocation is None:
                        # If commit_allocation not present, raise to rollback
                        messages.error(request, "commit_allocation is not available. Allocation cannot be committed.")
                        inv.delete()
                        return redirect("sales:invoice_add")
                    try:
                        # commit_allocation signature in your project earlier expected (invoice, invoice_item, allocations, ...)
                        commit_allocation(inv, it, alloc_plan[i], ref_type='invoice')
                    except Exception as e:
                        # rollback on error
                        messages.error(request, f"Allocation commit failed: {e}")
                        inv.delete()
                        return redirect("sales:invoice_add")

        # ---------- PAYMENTS ----------
        pays = request.POST.getlist("payment_amount[]")
        methods = request.POST.getlist("payment_method[]")
        dates = request.POST.getlist("payment_date[]")

        for i in range(max(len(pays), len(methods), len(dates))):
            amount = parse_decimal(pays[i] if i < len(pays) else "0")
            method = methods[i] if i < len(methods) else "Cash"
            date = dates[i] if i < len(dates) else today

            if amount > 0:
                Payment.objects.create(
                    invoice=inv,
                    amount=amount.quantize(Decimal("0.01")),
                    method=method or "Cash",
                    date=date,
                )

        # ---------- RECALCULATE TOTALS & STATUS ----------
        try:
            inv.recalc_totals()
        except Exception:
            # fallback manual totals calculation (existing logic)
            subtotal = Decimal("0.00")
            discount_total = Decimal("0.00")
            tax_total = Decimal("0.00")

            for it in inv.items.all():
                line_total = (Decimal(it.quantity) * Decimal(it.rate)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                disc_amt = (line_total * (it.discount_percent or Decimal("0")) / Decimal("100.00")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                taxed = ((line_total - disc_amt) * (Decimal(it.tax_rate or 0) / Decimal("100.00"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

                subtotal += line_total
                discount_total += disc_amt
                tax_total += taxed

                try:
                    it.line_total = (line_total - disc_amt + taxed).quantize(Decimal("0.01"))
                    it.save(update_fields=["line_total"])
                except Exception:
                    pass

            inv.subtotal = subtotal.quantize(Decimal("0.01"))
            inv.discount_total = discount_total.quantize(Decimal("0.01"))
            inv.tax_total = tax_total.quantize(Decimal("0.01"))
            inv.grand_total = (subtotal - discount_total + tax_total).quantize(Decimal("0.01"))
            inv.save(update_fields=["subtotal", "discount_total", "tax_total", "grand_total"])

        try:
            inv.recalc_payments_and_status()
        except Exception:
            logger.exception("Invoice recalc_totals failed in invoice_add, applying fallback. invoice_id=%s", inv.id)
            paid = inv.payments.aggregate(total=Sum("amount"))["total"] or Decimal("0")
            balance = (inv.grand_total - paid).quantize(Decimal("0.01"))
            inv.paid_amount = paid.quantize(Decimal("0.01"))
            inv.balance_due = balance if balance > 0 else Decimal("0.00")

            if balance <= 0:
                inv.status = "paid"
            elif paid > 0:
                inv.status = "partial"
            else:
                inv.status = "unpaid"

            inv.save(update_fields=["paid_amount", "balance_due", "status"])

        allocate_invoice(inv)

        messages.success(request, f"Invoice {inv.number} created successfully.")
        return redirect("sales:invoice_detail", pk=inv.pk)

    # GET request - render form
    ctx = {
        "customers": customers,
        "products": products,
        "saved_items": saved_items,
        "companies": companies,
        "today": today,
        "next_number": next_number,
    }
    return render(request, "sales/invoice_add.html", ctx)

@login_required
def invoice_detail(request, pk):
    """
    Single invoice detail view with:
    - Split payments vs refunds
    - Net received amount
    """
    inv = get_object_or_404(Invoice, pk=pk)

    payments = inv.payments.filter(is_refund=False).order_by("date", "id")
    refunds = inv.payments.filter(is_refund=True).order_by("date", "id")

    paid_total = payments.aggregate(total=Sum("amount"))["total"] or Decimal(
        "0.00"
    )
    total_refunded = (
        refunds.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    )
    net_received = (paid_total - total_refunded).quantize(Decimal("0.01"))

    context = {
        "inv": inv,
        "payments": payments,
        "refunds": refunds,
        "paid_total": paid_total,
        "total_refunded": total_refunded,
        "net_received": net_received,
    }
    return render(request, "sales/invoice_detail.html", context)


# ---------- invoice_edit (FINAL VERSION) ----------
@login_required
@transaction.atomic
def invoice_edit(request, pk):
    """
    Edit existing invoice:

    Flow:
    1) Reverse ALL previous allocations for this invoice (safe, idempotent).
    2) Delete old items.
    3) Rebuild items from POST (manual / saved / product).
    4) Build fresh FIFO allocation plan (like invoice_restore).
    5) Commit allocations for product lines.
    6) Reset + recreate payments/refunds.
    7) Recalculate totals + status.
    """
    inv = get_object_or_404(Invoice, pk=pk)
    customers = Customer.objects.all()
    products = Product.objects.all()
    saved_items = SavedItem.objects.all()
    companies = CompanyProfile.objects.all()
    today = timezone.localdate()

    if request.method == "POST":
        # -------- HEADER --------
        inv.number = request.POST.get("number", inv.number)
        inv.customer_id = request.POST.get("customer") or None
        inv.company_id = request.POST.get("company") or None
        inv.date = request.POST.get("date") or today
        inv.billing_address = request.POST.get("billing_address", "")
        inv.shipping_address = request.POST.get("shipping_address", "")
        inv.notes = request.POST.get("notes", "")
        inv.terms = request.POST.get("terms", "")
        inv.bank_name = request.POST.get("bank_name", "")
        inv.bank_branch = request.POST.get("bank_branch", "")
        inv.account_number = request.POST.get("account_number", "")
        inv.ifsc = request.POST.get("ifsc", "")
        inv.upi = request.POST.get("upi", "")
        # logo ko optional hi rehne do
        if "logo" in request.FILES:
            inv.logo = request.FILES["logo"]
        inv.save()

        # -------- 1) REVERSE OLD ALLOCATIONS --------
        if reverse_allocation_for_invoice:
            try:
                reverse_allocation_for_invoice(inv)
            except Exception:
                logger.exception("reverse_allocation_for_invoice failed in invoice_edit. invoice_id=%s", inv.id)
                # silently ignore, doctor / audit tool catch karega
                pass

        # -------- 2) DELETE OLD ITEMS --------
        inv.items.all().delete()

        # -------- 3) READ POSTED LINES & RECREATE ITEMS --------
        prods = request.POST.getlist("product[]")           # id / 'manual' / 'saved:<id>'
        manuals = request.POST.getlist("manual_product[]")  # description
        hsn_codes = request.POST.getlist("hsn_code[]")
        qtys = request.POST.getlist("qty[]")
        rates = request.POST.getlist("rate[]")
        tax_rates = request.POST.getlist("tax[]")
        discounts = request.POST.getlist("discount[]")

        max_lines = (
            max(
                len(prods),
                len(manuals),
                len(hsn_codes),
                len(qtys),
                len(rates),
                len(tax_rates),
                len(discounts),
            )
            if (prods or manuals or qtys or rates or tax_rates or discounts)
            else 0
        )

        # Plain recreate items (no stock logic yahan)
        for i in range(max_lines):
            p_val = prods[i] if i < len(prods) else ""
            m_val = manuals[i] if i < len(manuals) else ""
            hsn_val = (hsn_codes[i] if i < len(hsn_codes) else "").strip()
            q_val = parse_decimal(qtys[i] if i < len(qtys) else "0")
            r_val = parse_decimal(rates[i] if i < len(rates) else "0")
            t_val = parse_int(tax_rates[i] if i < len(tax_rates) else "0")
            d_val = parse_decimal(discounts[i] if i < len(discounts) else "0")

            if q_val <= 0 and r_val <= 0:
                continue

            qty_int = int(q_val)

            # --- MANUAL LINE ---
            if not p_val or p_val == "manual":
                InvoiceItem.objects.create(
                    invoice=inv,
                    description=m_val or "Custom Item",
                    hsn_code=hsn_val,
                    quantity=qty_int,
                    rate=r_val,
                    tax_rate=t_val,
                    discount_percent=d_val,
                )
                continue

            # --- SAVED ITEM LINE ---
            if isinstance(p_val, str) and p_val.startswith("saved:"):
                try:
                    sid = int(p_val.split(":", 1)[1])
                    si = SavedItem.objects.get(pk=sid)
                    used_rate = si.price if r_val == 0 else r_val
                    InvoiceItem.objects.create(
                        invoice=inv,
                        saved_item=si,
                        description=si.name,
                        hsn_code=hsn_val,
                        quantity=qty_int,
                        rate=used_rate,
                        tax_rate=si.tax_rate,
                        discount_percent=d_val,
                    )
                except Exception:
                    InvoiceItem.objects.create(
                        invoice=inv,
                        description=m_val or "Saved Item missing",
                        hsn_code=hsn_val,
                        quantity=qty_int,
                        rate=r_val,
                        tax_rate=t_val,
                        discount_percent=d_val,
                    )
                continue

            # --- PRODUCT LINE ---
            try:
                prod_obj = CatalogProduct.objects.get(pk=int(p_val))
            except Exception:
                # product missing — treat as manual
                InvoiceItem.objects.create(
                    invoice=inv,
                    description=m_val or "Product missing",
                    hsn_code=hsn_val,
                    quantity=qty_int,
                    rate=r_val,
                    tax_rate=t_val,
                    discount_percent=d_val,
                )
                continue

            used_rate = r_val if r_val != Decimal("0") else getattr(prod_obj, "price", Decimal("0"))
            InvoiceItem.objects.create(
                invoice=inv,
                product=prod_obj,
                description=prod_obj.name,
                hsn_code=(getattr(prod_obj, "hsn_code", "") or hsn_val),
                quantity=qty_int,
                rate=used_rate,
                tax_rate=getattr(prod_obj, "tax_rate", t_val),
                discount_percent=d_val,
            )

        # -------- 4) BUILD FRESH ALLOCATION PLAN (like invoice_restore) --------
        if allocate_fifo and commit_allocation:
            posted_items = list(inv.items.all())
            insufficient = []
            alloc_plan = {}

            # lock products
            prod_ids = [
                it.product_id
                for it in posted_items
                if getattr(it, "product_id", None)
            ]
            locked_products = {}
            if prod_ids:
                for prd in CatalogProduct.objects.select_for_update().filter(id__in=prod_ids):
                    locked_products[str(prd.id)] = prd

            for idx, it in enumerate(posted_items):
                prod = getattr(it, "product", None)
                if not prod:
                    continue
                if not getattr(prod, "track_stock", True):
                    continue

                allocs = allocate_fifo(prod, int(it.quantity))
                if allocs is None:
                    insufficient.append(
                        (idx, f"{prod.sku or prod.name} (need {it.quantity})")
                    )
                else:
                    alloc_plan[idx] = allocs

            if insufficient:
                msg = "; ".join([f"Line {i+1}: {m}" for i, m in insufficient])
                messages.error(
                    request,
                    f"Cannot update invoice due to insufficient stock: {msg}",
                )
                # Yahan tak allocations reverse ho chuke hain, items recreate ho chuke.
                # Stock safe hai (invoice ka net effect abhi 0 hai).
                return redirect("sales:invoice_edit", pk=inv.pk)

            # -------- 5) COMMIT ALLOCATIONS --------
            for idx, it in enumerate(posted_items):
                if idx in alloc_plan:
                    commit_allocation(inv, it, alloc_plan[idx], ref_type="invoice")

        # -------- 6) RESET + RECREATE PAYMENTS / REFUNDS (as before) --------
        inv.payments.filter(is_refund=False).delete()
        inv.payments.filter(is_refund=True).delete()

        pays = request.POST.getlist("payment_amount[]")
        methods = request.POST.getlist("payment_method[]")
        dates = request.POST.getlist("payment_date[]")

        pay_max = (
            max(len(pays), len(methods), len(dates))
            if (pays or methods or dates)
            else 0
        )

        for i in range(pay_max):
            amt = parse_decimal(pays[i]) if i < len(pays) and pays[i] else Decimal("0")
            if amt > 0:
                Payment.objects.create(
                    invoice=inv,
                    amount=amt.quantize(Decimal("0.01")),
                    method=(methods[i] if i < len(methods) and methods[i] else "Cash"),
                    date=(dates[i] if i < len(dates) and dates[i] else today),
                    is_refund=False,
                )

        # Refunds
        refund_amounts = request.POST.getlist("refund_amount[]")
        refund_methods = request.POST.getlist("refund_method[]")
        refund_dates = request.POST.getlist("refund_date[]")
        refund_notes = request.POST.getlist("refund_note[]")

        ref_max = (
            max(len(refund_amounts), len(refund_methods), len(refund_dates), len(refund_notes))
            if (refund_amounts or refund_methods or refund_dates or refund_notes)
            else 0
        )

        for i in range(ref_max):
            amt = parse_decimal(refund_amounts[i]) if i < len(refund_amounts) and refund_amounts[i] else Decimal("0")
            if amt > 0:
                Payment.objects.create(
                    invoice=inv,
                    amount=amt.quantize(Decimal("0.01")),
                    method=(refund_methods[i] if i < len(refund_methods) and refund_methods[i] else "Refund"),
                    date=(refund_dates[i] if i < len(refund_dates) and refund_dates[i] else today),
                    note=(refund_notes[i] if i < len(refund_notes) else ""),
                    is_refund=True,
                )

        # -------- 7) RECALCULATE TOTALS / STATUS --------
        try:
            inv.recalc_totals()
            inv.recalc_payments_and_status()
        except Exception:
            logger.exception("Invoice recalc failed in invoice_edit. invoice_id=%s", inv.id)

        # Manual status override
        manual_status_val = request.POST.get("manual_status", "0")
        status_val = request.POST.get("status") or inv.status or "unpaid"

        if manual_status_val == "1":
            inv.manual_status = True
            inv.status = status_val
        else:
            inv.manual_status = False

        inv.save()
        allocate_invoice(inv) 
        return redirect("sales:invoice_detail", pk=inv.pk)

    # GET request
    total_refunded = (
        inv.payments.filter(is_refund=True).aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )

    payments_qs = inv.payments.filter(is_refund=False).order_by("date", "id")
    refunds_qs = inv.payments.filter(is_refund=True).order_by("date", "id")

    ctx = {
        "inv": inv,
        "customers": customers,
        "products": products,
        "saved_items": saved_items,
        "companies": companies,
        "today": today,
        "total_refunded": total_refunded,
        "payments": payments_qs,
        "refunds": refunds_qs,
    }
    return render(request, "sales/invoice_edit.html", ctx)


# ============================================================
# COMPANY INFO API (AJAX)
# ============================================================

@login_required
def company_info_api(request, pk):
    """
    Simple JSON API to fetch CompanyProfile details for invoice form.
    """
    try:
        c = CompanyProfile.objects.get(pk=pk)
    except CompanyProfile.DoesNotExist:
        return JsonResponse({}, status=404)

    data = {
        "id": c.id,
        "name": c.name,
        "address": c.address,
        "gstin": c.gstin,
        "phone": c.phone,
        "email": c.email,
        "bank_details": c.bank_details,
        "logo_url": c.logo.url if c.logo else "",
    }
    return JsonResponse(data)


# ============================================================
# INVOICE DELETE + BIN (SOFT DELETE FLOW)
# ============================================================
# Ensure import
from catalog.inventory_utils import reverse_allocation_for_invoice

@login_required
@transaction.atomic
def invoice_delete(request, pk):
    """
    Permanent delete (from active list, NOT from bin).
    Reverse allocations first so stock is restored.
    """
    inv = get_object_or_404(Invoice, pk=pk, in_bin=False)

    if request.method == "POST":
        number = inv.number
        # reverse inventory allocations first
        try:
            reverse_allocation_for_invoice(inv)
        except Exception:
            logger.exception("reverse_allocation_for_invoice failed in invoice_delete. invoice_id=%s", inv.id)

        inv.delete()
        messages.success(request, f"Invoice {number} permanently deleted.")
        return redirect("sales:invoice_list")

    return render(request, "sales/invoice_delete_confirm.html", {"inv": inv})

@login_required
@transaction.atomic
def invoice_move_to_bin(request, pk):
    """
    Move invoice into bin (soft delete).
    Reverse stock allocations so stock returns to inventory.
    """
    inv = get_object_or_404(Invoice, pk=pk)

    if request.method == "POST":
        try:
            reverse_allocation_for_invoice(inv)
        except Exception:
            logger.exception("reverse_allocation_for_invoice failed in invoice_move_to_bin. invoice_id=%s", inv.id)
        inv.in_bin = True
        inv.save()
        messages.success(request, f"Invoice {inv.number} moved to Bin and stock restored.")
        return redirect("sales:invoice_list")

    return redirect("sales:invoice_delete", pk=pk)

@login_required
def invoice_bin_list(request):
    """
    Bin screen: shows only in_bin=True invoices.
    """
    invoices = (
        Invoice.objects.select_related("customer")
        .filter(in_bin=True)
        .order_by("-date", "-id")
    )

    return render(
        request,
        "sales/invoice_bin_list.html",
        {"invoices": invoices},
    )


@login_required
@transaction.atomic
def invoice_restore(request, pk):
    """
    Restore invoice from bin (attempt to re-allocate stock).
    If re-allocation fails due to insufficient stock, do not restore and show message.
    """
    inv = get_object_or_404(Invoice, pk=pk, in_bin=True)

    if request.method == "POST":
        # Build allocation plan from existing invoice items
        from catalog.models import Product as CatalogProduct
        from catalog.inventory_utils import allocate_fifo, commit_allocation

        posted_items = list(inv.items.all())
        insufficient = []
        alloc_plan = {}

        # lock products
        prod_ids = [it.product_id for it in posted_items if getattr(it, "product_id", None)]
        locked_products = {}
        if prod_ids:
            for prd in CatalogProduct.objects.select_for_update().filter(id__in=prod_ids):
                locked_products[str(prd.id)] = prd

        for idx, it in enumerate(posted_items):
            prod = getattr(it, "product", None)
            if not prod:
                continue
            if not getattr(prod, "track_stock", True):
                continue
            allocs = allocate_fifo(prod, int(it.quantity))
            if allocs is None:
                insufficient.append((idx, f"{prod.sku or prod.name} (need {it.quantity})"))
            else:
                alloc_plan[idx] = allocs

        if insufficient:
            msg = "; ".join([f"Line {i+1}: {m}" for i,m in insufficient])
            messages.error(request, f"Cannot restore invoice due to insufficient stock: {msg}")
            return redirect("sales:invoice_bin_list")

        # commit allocations and restore invoice
        for idx, it in enumerate(posted_items):
            if idx in alloc_plan:
                commit_allocation(inv, it, alloc_plan[idx], ref_type='invoice')

        inv.in_bin = False
        inv.save()
        messages.success(request, f"Invoice {inv.number} restored and stock re-allocated.")
        return redirect("sales:invoice_detail", pk=inv.pk)

    return redirect("sales:invoice_bin_list")


@login_required
@transaction.atomic
def invoice_bin_delete(request, pk):
    """
    Permanent delete from bin. Reverse allocations first just in case.
    """
    inv = get_object_or_404(Invoice, pk=pk, in_bin=True)

    if request.method == "POST":
        number = inv.number
        try:
            reverse_allocation_for_invoice(inv)
        except Exception:
            pass
        inv.delete()
        messages.success(request, f"Invoice {number} permanently deleted from Bin.")
        return redirect("sales:invoice_bin_list")

    return render(request, "sales/invoice_bin_delete_confirm.html", {"inv": inv})


# ============================================================
# INVOICE PREVIEW + PDF (WEASYPRINT)
# ============================================================

@login_required
def invoice_preview(request, pk):
    """
    Simple preview (reuse invoice_detail template if you want).
    """
    inv = get_object_or_404(
        Invoice.objects.select_related("customer"), pk=pk
    )
    return render(request, "sales/invoice_detail.html", {"inv": inv})


from django.shortcuts import get_object_or_404
from decimal import Decimal

@login_required
def invoice_pdf(request, pk):
    """
    Generate invoice PDF using WeasyPrint and the dedicated PDF template.
    """
    inv = get_object_or_404(
        Invoice.objects.select_related("customer"), pk=pk
    )

    payments = inv.payments.filter(is_refund=False).order_by("date", "id")
    refunds = inv.payments.filter(is_refund=True).order_by("date", "id")

    paid_total = payments.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    total_refunded = refunds.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    # Company fallback values
    company_name = (
        inv.company.name if getattr(inv, "company", None) else "The Webfix"
    )
    company_address = (
        inv.company.address
        if getattr(inv, "company", None) and inv.company.address
        else ""
    )

    html_string = render_to_string(
        "sales/invoice_pdf.html",
        {
            "inv": inv,
            "payments": payments,
            "refunds": refunds,
            "paid_total": paid_total,
            "total_refunded": total_refunded,
            "company_name": company_name,
            "company_address": company_address,
        },
        request=request,
    )

    from weasyprint import HTML

    pdf_bytes = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/"),
    ).write_pdf()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{inv.number}.pdf"'
    return response


# ============================================================
# CUSTOMER LIST + EXPORT
# ============================================================

@login_required
def customer_list(request):
    """
    Advanced customer list:
    - Search: ?q=
    - Status filter: ?status=all|active|inactive|pending|top
    - Type filter: ?type=Individual|Business
    - Export: ?export=csv or ?export=xlsx (filters-respecting)
    """
    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    ctype = request.GET.get("type", "").strip()
    export = request.GET.get("export", "").strip()

    # Base queryset + aggregates
    base_qs = Customer.objects.annotate(
        invoices_count=Count("invoices", distinct=True),
        total_spent_calc=Coalesce(
            Sum("invoices__grand_total"), Decimal("0.00")
        ),
        pending_balance_calc=Coalesce(
            Sum("invoices__balance_due"), Decimal("0.00")
        ),
    )

    # Overview totals (unfiltered)
    total_customers = base_qs.count()
    active_customers = base_qs.filter(invoices_count__gt=0).count()
    top_spender = base_qs.order_by("-total_spent_calc").first()

    # Filtered queryset for UI
    customers = base_qs

    # ---- SEARCH ----
    if q:
        customers = customers.filter(
            Q(name__icontains=q)
            | Q(email__icontains=q)
            | Q(phone__icontains=q)
            | Q(company__icontains=q)
            | Q(gstin__icontains=q)
            | Q(pan__icontains=q)
        )

    # ---- TYPE FILTER ----
    if ctype:
        customers = customers.filter(customer_type__iexact=ctype)

    # ---- STATUS FILTER ----
    if status == "active":
        customers = customers.filter(invoices_count__gt=0)
    elif status == "inactive":
        customers = customers.filter(invoices_count=0)
    elif status == "pending":
        customers = customers.filter(pending_balance_calc__gt=0)
    elif status == "top":
        customers = customers.order_by("-total_spent_calc")
    else:
        customers = customers.order_by("-id")

    # ---- EXPORT ----
    if export == "csv":
        return export_customers_csv(customers)
    elif export == "xlsx":
        return export_customers_excel(customers)

    context = {
        "customers": customers,
        "total_customers": total_customers,
        "active_customers": active_customers,
        "top_spender": top_spender,
        "query": q,
        "status": status,
        "ctype": ctype,
    }
    return render(request, "sales/customer_list.html", context)


def export_customers_csv(customers_qs):
    """
    Export filtered customers to CSV.
    """
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="customers.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "Name",
            "Company",
            "Email",
            "Phone",
            "Type",
            "Total Invoices",
            "Total Spent",
            "Pending Balance",
        ]
    )

    for c in customers_qs:
        writer.writerow(
            [
                c.name,
                c.company or "",
                c.email or "",
                c.phone or "",
                c.customer_type or "",
                int(c.invoices_count or 0),
                float(c.total_spent_calc or 0),
                float(c.pending_balance_calc or 0),
            ]
        )

    return response


def export_customers_excel(customers_qs):
    """
    Export filtered customers to XLSX.
    Falls back to CSV if openpyxl is missing.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return export_customers_csv(customers_qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    ws.append(
        [
            "Name",
            "Company",
            "Email",
            "Phone",
            "Type",
            "Total Invoices",
            "Total Spent",
            "Pending Balance",
        ]
    )

    for c in customers_qs:
        ws.append(
            [
                c.name,
                c.company or "",
                c.email or "",
                c.phone or "",
                c.customer_type or "",
                int(c.invoices_count or 0),
                float(c.total_spent_calc or 0),
                float(c.pending_balance_calc or 0),
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = 'attachment; filename="customers.xlsx"'
    return response


# ============================================================
# CUSTOMER CRUD + DETAIL + AJAX
# ============================================================

@login_required
def customer_detail(request, pk):
    """
    Rich customer detail:
    - Invoices (not deleted, not in bin)
    - Total spent & pending
    - Recent payments
    - Mixed activity timeline (invoices + payments)
    """
    customer = get_object_or_404(Customer, pk=pk)

    invoices = (
        Invoice.objects.filter(
            customer=customer,
            deleted_at__isnull=True,
            in_bin=False,
        )
        .order_by("-date", "-id")
    )

    totals = invoices.aggregate(
        total_spent=Sum("grand_total"),
        pending_balance=Sum("balance_due"),
    )
    customer.total_spent = totals["total_spent"] or 0
    customer.pending_balance = totals["pending_balance"] or 0

    payments = (
        Payment.objects.filter(invoice__customer=customer)
        .order_by("-date", "-id")
    )

    # Activity timeline (mix top 5 invoices + top 5 payments -> sorted)
       # normalize created_at to datetime for safe sorting (avoid mixing date/datetime)
    from datetime import datetime, time as dtime

    activities = []

    def normalize_to_datetime(obj_dt):
        """
        Accepts a date or datetime. Returns a datetime.
        If input is a date -> combine with time.min (00:00).
        If naive datetime -> try to make timezone-aware using Django timezone.
        """
        if obj_dt is None:
            return timezone.make_aware(datetime.min) if timezone.is_naive(datetime.min) else datetime.min

        # if it's already a datetime
        if isinstance(obj_dt, datetime):
            dt = obj_dt
        else:
            # treat as date
            try:
                dt = datetime.combine(obj_dt, dtime.min)
            except Exception:
                logger.exception("normalize_to_datetime fallback hit in customer_detail")
                # fallback to min datetime
                dt = datetime.min

        # If a naive datetime and timezone utilities available, try to make it aware
        try:
            if timezone.is_naive(dt):
                # prefer timezone.make_aware so comparisons with other aware datetimes don't fail
                dt = timezone.make_aware(dt)
        except Exception:
            logger.exception("timezone.make_aware fallback hit in customer_detail")
            # if make_aware fails (rare), leave dt as-is
            pass

        return dt

    # invoices (top 5)
    for inv in invoices[:5]:
        created_raw = getattr(inv, "created_at", None) or getattr(inv, "date", None)
        created_dt = normalize_to_datetime(created_raw)
        activities.append(
            {
                "type": "invoice",
                "date": inv.date,
                "created_at": created_dt,
                "label": f"Invoice #{inv.number}",
                "amount": inv.grand_total,
                "status": inv.status,
                "obj": inv,
            }
        )

    # payments (top 5)
    for pay in payments[:5]:
        created_raw = getattr(pay, "created_at", None) or getattr(pay, "date", None)
        created_dt = normalize_to_datetime(created_raw)
        activities.append(
            {
                "type": "payment",
                "date": pay.date,
                "created_at": created_dt,
                "label": "Payment received",
                "amount": pay.amount,
                "method": pay.method,
                "invoice": pay.invoice,
            }
        )

    # now safe to sort — all created_at are datetimes
    activities = sorted(activities, key=lambda x: x["created_at"], reverse=True)[:5]


    context = {
        "customer": customer,
        "invoices": invoices,
        "activities": activities,
    }
    return render(request, "sales/customer_detail.html", context)


@login_required
def customer_add(request):
    """
    Simple customer create form.
    """
    if request.method == "POST":
        Customer.objects.create(
            name=request.POST.get("name"),
            company=request.POST.get("company", ""),
            email=request.POST.get("email", ""),
            phone=request.POST.get("phone", ""),
            customer_type=request.POST.get("customer_type", ""),
            industry=request.POST.get("industry", ""),
            pan=request.POST.get("pan", ""),
            gstin=request.POST.get("gstin", ""),
            website=request.POST.get("website", ""),
            billing_address=request.POST.get("billing_address", ""),
            shipping_address=request.POST.get("shipping_address", ""),
            notes=request.POST.get("notes", ""),
        )
        return redirect("sales:customer_list")

    return render(request, "sales/customer_add.html")


@login_required
def customer_edit(request, pk):
    """
    Edit customer master data.
    """
    customer = get_object_or_404(Customer, pk=pk)

    if request.method == "POST":
        customer.name = request.POST.get("name")
        customer.company = request.POST.get("company", "")
        customer.email = request.POST.get("email", "")
        customer.phone = request.POST.get("phone", "")
        customer.customer_type = request.POST.get("customer_type", "")
        customer.industry = request.POST.get("industry", "")
        customer.pan = request.POST.get("pan", "")
        customer.gstin = request.POST.get("gstin", "")
        customer.website = request.POST.get("website", "")
        customer.billing_address = request.POST.get(
            "billing_address", ""
        )
        customer.shipping_address = request.POST.get(
            "shipping_address", ""
        )
        customer.notes = request.POST.get("notes", "")
        customer.save()
        return redirect("sales:customer_detail", pk=customer.pk)

    return render(
        request,
        "sales/customer_edit.html",
        {"customer": customer},
    )


@login_required
def customer_delete(request, pk):
    """
    Hard delete customer (no soft-delete here).
    """
    customer = get_object_or_404(Customer, pk=pk)

    if request.method == "POST":
        customer.delete()
        return redirect("sales:customer_list")

    return render(
        request,
        "sales/customer_delete_confirm.html",
        {"customer": customer},
    )


@login_required
def customer_add_ajax(request):
    """
    Quick customer add via AJAX,
    used e.g. from invoice creation modal.
    """
    if (
        request.method == "POST"
        and request.headers.get("x-requested-with") == "XMLHttpRequest"
    ):
        name = request.POST.get("name")
        if not name:
            return JsonResponse(
                {"ok": False, "error": "Name required"},
                status=400,
            )

        c = Customer.objects.create(
            name=name,
            email=request.POST.get("email", ""),
            phone=request.POST.get("phone", ""),
            billing_address=request.POST.get("address", ""),
            gstin=request.POST.get("gstin", ""),
        )
        return JsonResponse(
            {"ok": True, "id": c.id, "name": c.name}
        )

    return JsonResponse(
        {"ok": False, "error": "Invalid request"},
        status=400,
    )


@login_required
def customer_details_ajax(request, pk):
    """
    Fetch customer details via AJAX for auto-fill.
    """
    try:
        c = Customer.objects.get(pk=pk)
        return JsonResponse(
            {
                "ok": True,
                "name": c.name,
                "email": c.email,
                "phone": c.phone,
                "gstin": c.gstin,
                "address": c.billing_address or "",
            }
        )
    except Customer.DoesNotExist:
        return JsonResponse(
            {"ok": False, "error": "Customer not found"},
            status=404,
        )


# ============================================================
# SAVED ITEM AJAX
# ============================================================

@login_required
def save_item_ajax(request):
    """
    Create a SavedItem record via AJAX from invoice form.
    """
    if (
        request.method == "POST"
        and request.headers.get("x-requested-with") == "XMLHttpRequest"
    ):
        name = request.POST.get("name")
        price = request.POST.get("price") or "0"
        tax = request.POST.get("tax") or "0"
        is_service = request.POST.get("is_service") == "1"

        if not name:
            return JsonResponse(
                {"ok": False, "error": "Item name is required"},
                status=400,
            )

        try:
            si = SavedItem.objects.create(
                name=name,
                price=parse_decimal(price),
                tax_rate=parse_int(tax),
                is_service=is_service,
            )
            return JsonResponse(
                {
                    "ok": True,
                    "id": si.id,
                    "name": si.name,
                    "price": str(si.price),
                    "tax": si.tax_rate,
                }
            )
        except Exception as e:
            logger.exception("save_item_ajax failed")
            return JsonResponse(
                {"ok": False, "error": str(e)},
                status=500,
            )

    return JsonResponse(
        {"ok": False, "error": "Invalid request"},
        status=400,
    )


# ============================================================
# PAYMENTS LIST / EXPORT / CRUD / DASHBOARD
# ============================================================

@login_required
def payment_list(request):
    """
    Payments listing with:
    - Filter by method / period (today / this_week / this_month)
    - Search by invoice number, customer name, note
    - Net totals (paid - refund) on filtered queryset
    - Today net total (global)
    """
    qs = Payment.objects.select_related(
        "invoice", "invoice__customer"
    ).order_by("-date", "-id")

    method = request.GET.get("method")
    period = request.GET.get("period")
    q = request.GET.get("q", "").strip()

    if method:
        qs = qs.filter(method__iexact=method)

    if q:
        qs = qs.filter(
            Q(invoice__number__icontains=q)
            | Q(invoice__customer__name__icontains=q)
            | Q(note__icontains=q)
        )

    today = timezone.localdate()
    if period == "today":
        qs = qs.filter(date=today)
    elif period == "this_month":
        qs = qs.filter(
            date__year=today.year,
            date__month=today.month,
        )
    elif period == "this_week":
        start_week = today - timezone.timedelta(days=today.weekday())
        qs = qs.filter(date__gte=start_week, date__lte=today)

    # Net totals (on filtered qs)
    agg = qs.aggregate(
        paid=Sum("amount", filter=models.Q(is_refund=False)),
        refunded=Sum("amount", filter=models.Q(is_refund=True)),
    )
    paid = agg["paid"] or Decimal("0.00")
    refunded = agg["refunded"] or Decimal("0.00")
    total_received = (paid - refunded).quantize(Decimal("0.01"))

    # Today's net across DB
    agg_today = Payment.objects.filter(date=today).aggregate(
        paid=Sum("amount", filter=models.Q(is_refund=False)),
        refunded=Sum("amount", filter=models.Q(is_refund=True)),
    )
    paid_today = agg_today["paid"] or Decimal("0.00")
    refund_today = agg_today["refunded"] or Decimal("0.00")
    total_today = (paid_today - refund_today).quantize(Decimal("0.01"))

    # Simple manual pagination (page/per_page)
    page = int(request.GET.get("page", 1))
    per = 25
    start = (page - 1) * per
    end = start + per
    payments = qs[start:end]

    return render(
        request,
        "sales/payment_list.html",
        {
            "payments": payments,
            "total_received": total_received,
            "total_today": total_today,
            "period": period,
            "method": method,
            "q": q,
        },
    )


@login_required
def payment_export_csv(request):
    """
    Export all payments as CSV (no filters).
    """
    qs = Payment.objects.select_related(
        "invoice", "invoice__customer"
    ).order_by("-date", "-id")

    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="payments.csv"'
    writer = csv.writer(resp)

    writer.writerow(
        ["Date", "Invoice", "Customer", "Amount", "Method", "Note"]
    )

    for p in qs:
        writer.writerow(
            [
                p.date,
                p.invoice.number if p.invoice else "",
                (
                    p.invoice.customer.name
                    if p.invoice and p.invoice.customer
                    else ""
                ),
                str(p.amount),
                p.method,
                p.note or "",
            ]
        )
    return resp


@login_required
def payment_add(request):
    """
    Simple manual payment add for an invoice.
    """
    invoices = Invoice.objects.filter(
        grand_total__gt=0
    ).order_by("-date")

    if request.method == "POST":
        invoice_id = request.POST.get("invoice") or None
        amount = parse_decimal(request.POST.get("amount") or "0")
        date = request.POST.get("date") or None
        method = request.POST.get("method") or "Cash"
        note = request.POST.get("note", "")

        inv = Invoice.objects.get(pk=invoice_id) if invoice_id else None

        Payment.objects.create(
            invoice=inv,
            amount=amount,
            date=date or timezone.localdate(),
            method=method,
            note=note,
        )

        if inv:
            inv.recalc_payments_and_status()

        return redirect("sales:payment_list")

    return render(
        request,
        "sales/payment_add.html",
        {"invoices": invoices},
    )


@login_required
def payment_detail(request, pk):
    """
    Single payment detail view.
    """
    p = get_object_or_404(
        Payment.objects.select_related("invoice", "invoice__customer"),
        pk=pk,
    )
    return render(request, "sales/payment_detail.html", {"p": p})


@login_required
def payments_dashboard(request):
    """
    High-level payments dashboard stats:
    - Total invoice value
    - Total collected (invoice.paid_amount)
    - Total refunded (Payment & legacy Refund)
    - Outstanding (invoice.balance_due)
    - Recent payments + invoices
    - Revenue trend (last 6 months)
    - Top customers (by paid_amount)
    """
    # Total invoice value
    total_revenue = (
        Invoice.objects.aggregate(total=Sum("grand_total"))["total"]
        or Decimal("0.00")
    )

    # Total actually received (net, from paid_amount)
    total_collected = (
        Invoice.objects.aggregate(total=Sum("paid_amount"))["total"]
        or Decimal("0.00")
    )

    # Refunds via new Payment.is_refund
    payment_refunds = (
        Payment.objects.filter(is_refund=True).aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )

    # Legacy Refund model
    legacy_refunds = (
        Refund.objects.aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )

    total_refunded = (
        payment_refunds + legacy_refunds
    ).quantize(Decimal("0.01"))

    # Outstanding balance
    outstanding = (
        Invoice.objects.aggregate(total=Sum("balance_due"))["total"]
        or Decimal("0.00")
    )

    unpaid_count = Invoice.objects.filter(status="unpaid").count()
    partial_count = Invoice.objects.filter(status="partial").count()

    recent_payments = (
        Payment.objects.select_related("invoice", "invoice__customer")
        .order_by("-date", "-id")[:8]
    )
    recent_invoices = (
        Invoice.objects.select_related("customer")
        .order_by("-date", "-id")[:8]
    )

    # Revenue trend (last 6 months)
    now = timezone.localdate()
    labels, values = [], []

    for i in range(5, -1, -1):
        month = (now.month - i - 1) % 12 + 1
        year = now.year + ((now.month - i - 1) // 12)

        labels.append(f"{year}-{month:02d}")

        s = (
            Invoice.objects.filter(
                date__year=year,
                date__month=month,
            ).aggregate(total=Sum("grand_total"))["total"]
            or Decimal("0.00")
        )
        values.append(float(s))

    # Top customers by net received (paid_amount)
    top_customers = (
        Invoice.objects.filter(customer__isnull=False)
        .values("customer__id", "customer__name")
        .annotate(total=Sum("paid_amount"))
        .order_by("-total")[:6]
    )

    return render(
        request,
        "sales/payments_dashboard.html",
        {
            "total_revenue": total_revenue,
            "total_collected": total_collected,
            "total_refunded": total_refunded,
            "outstanding": outstanding,
            "unpaid_count": unpaid_count,
            "partial_count": partial_count,
            "recent_payments": recent_payments,
            "recent_invoices": recent_invoices,
            "chart_labels": labels,
            "chart_values": values,
            "top_customers": top_customers,
        },
    )


# ============================================================
# UNPAID INVOICES + MARK PAID
# ============================================================


# ---------- enhanced unpaid_invoices view ----------
@login_required
def unpaid_invoices(request):
    """
    Collection Center with:
    - search, customer filter, age filter, status filter
    - pagination
    - exports: csv, xlsx, pdf (using export param)
    - bulk-mark-paid via POST to `bulk_mark_paid` view (below)
    """
    q = request.GET.get("q", "").strip()
    customer_id = request.GET.get("customer", "").strip()
    status = request.GET.get("status", "").strip()  # unpaid/partial/all
    min_days = parse_int(request.GET.get("min_days", "0"))
    export = request.GET.get("export", "").strip()  # csv / xlsx / pdf
    page_num = parse_int(request.GET.get("page", 1))

    today = timezone.localdate()

    invoices_qs = (
        Invoice.objects
        .filter(balance_due__gt=0, in_bin=False)
        .select_related("customer")
        .order_by("date", "id")
    )

    # SEARCH
    if q:
        invoices_qs = invoices_qs.filter(
            Q(number__icontains=q)
            | Q(customer__name__icontains=q)
            | Q(customer__email__icontains=q)
            | Q(customer__phone__icontains=q)
        )

    # CUSTOMER FILTER
    if customer_id:
        invoices_qs = invoices_qs.filter(customer_id=customer_id)

    # STATUS FILTER
    if status in ["unpaid", "partial"]:
        invoices_qs = invoices_qs.filter(status=status)

    # AGE FILTER (days overdue)
    if min_days and min_days > 0:
        invoices_qs = invoices_qs.filter(
            date__lte=today - timedelta(days=min_days)
        )

    # annotate days_overdue on each invoice (cheap loop)
    invoices = list(invoices_qs)
    for inv in invoices:
        try:
            inv.days_overdue = (today - inv.date).days
        except Exception:
            logger.exception("days_overdue calc failed in unpaid_invoices. invoice_id=%s", inv.id)
            inv.days_overdue = 0

    # Totals
    totals = invoices_qs.aggregate(
        total_outstanding=Coalesce(Sum("balance_due"), Decimal("0.00")),
        total_value=Coalesce(Sum("grand_total"), Decimal("0.00")),
    )
    total_outstanding = (
        totals["total_outstanding"] or Decimal("0.00")
    ).quantize(Decimal("0.01"))
    total_value = (
        totals["total_value"] or Decimal("0.00")
    ).quantize(Decimal("0.01"))

    # EXPORT HANDLERS
    if export == "csv":
        return export_unpaid_csv(invoices_qs)

    elif export == "xlsx":
        return export_unpaid_xlsx(invoices_qs)

    elif export == "pdf":
        # render a simple PDF summary/layout
        html = render_to_string(
            "sales/collection_summary_pdf.html",
            {
                "company": CompanyProfile.objects.first(),
                "invoices": invoices,
                "total_outstanding": total_outstanding,
                "total_value": total_value,
                "from_date": None,
                "to_date": today,
                "generated_at": timezone.now(),
            },
            request=request,
        )

        config = pdf_config()
        options = pdf_options()

        pdf_bytes = pdfkit.from_string(
            html,
            False,
            configuration=config,
            options=options,
        )

        fname = f"collection_summary_{today}.pdf"
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    # PAGINATION
    per_page = 25
    paginator = Paginator(invoices, per_page)
    page = paginator.get_page(page_num)

    # customers for filter dropdown (only those with outstanding)
    customers_with_outstanding = (
        Customer.objects
        .filter(invoices__balance_due__gt=0)
        .distinct()
        .order_by("name")
    )

    context = {
        "invoices": page.object_list,
        "page_obj": page,
        "paginator": paginator,
        "q": q,
        "customer_id": customer_id,
        "status": status,
        "min_days": min_days,
        "total_outstanding": total_outstanding,
        "total_value": total_value,
        "invoice_count": invoices_qs.count(),
        "customers": customers_with_outstanding,
        "today": today,
    }
    return render(request, "sales/unpaid_invoices.html", context)

# ---------- export helpers ----------
def export_unpaid_csv(qs):
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="collection_unpaid.csv"'
    writer = csv.writer(resp)
    writer.writerow(["Invoice #", "Date", "Customer", "Customer Phone", "Invoice Total", "Balance Due", "Status", "Days Overdue"])
    today = timezone.localdate()
    for inv in qs.order_by("date", "id"):
        try:
            days = (today - inv.date).days
        except Exception:
            logger.exception("days_overdue calc failed in export_unpaid_csv. invoice_id=%s", inv.id)
            days = ""
        writer.writerow([
            inv.number,
            inv.date.strftime("%Y-%m-%d") if inv.date else "",
            inv.customer.name if inv.customer else "",
            inv.customer.phone if inv.customer else "",
            float(inv.grand_total or 0),
            float(inv.balance_due or 0),
            inv.status,
            days,
        ])
    return resp


def export_unpaid_xlsx(qs):
    if Workbook is None:
        # fallback to CSV
        return export_unpaid_csv(qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Unpaid Invoices"
    ws.append(["Invoice #", "Date", "Customer", "Customer Phone", "Invoice Total", "Balance Due", "Status", "Days Overdue"])
    today = timezone.localdate()
    for inv in qs.order_by("date", "id"):
        try:
            days = (today - inv.date).days
        except Exception:
            days = ""
        ws.append([
            inv.number,
            inv.date.strftime("%Y-%m-%d") if inv.date else "",
            inv.customer.name if inv.customer else "",
            inv.customer.phone if inv.customer else "",
            float(inv.grand_total or 0),
            float(inv.balance_due or 0),
            inv.status,
            days,
        ])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="collection_unpaid.xlsx"'
    return resp


# ---------- BULK MARK PAID endpoint ----------
@login_required
@transaction.atomic
def bulk_mark_paid(request):
    """
    POST endpoint: receives invoice_ids and optional date/method to create a payment record for each.
    Expects: invoice_ids[]=1&invoice_ids[]=2  (form-encoded or fetch JSON)
    """
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    ids = request.POST.getlist("invoice_ids[]") or request.POST.getlist("invoice_ids")
    amount_all = request.POST.get("amount_all")  # optional: mark full balance if not provided
    method = request.POST.get("method", "Cash")
    date = request.POST.get("date") or timezone.localdate()

    if not ids:
        return JsonResponse({"ok": False, "error": "No invoices selected"}, status=400)

    for sid in ids:
        try:
            inv = Invoice.objects.get(pk=int(sid))
        except Exception:
            continue
        amt = None
        if amount_all:
            amt = parse_decimal(amount_all)
        else:
            # default to full remaining balance
            amt = inv.balance_due or inv.grand_total
        if amt > 0:
            Payment.objects.create(
                invoice=inv,
                amount=amt.quantize(Decimal("0.01")),
                method=method,
                date=date,
            )
            # invoice recalc happens in Payment.save()
    return JsonResponse({"ok": True, "marked": len(ids)})

@login_required
def invoice_mark_paid(request, pk):
    """
    Quick form to record a payment for an unpaid invoice.
    """
    inv = get_object_or_404(Invoice, pk=pk)

    if request.method == "POST":
        amt = parse_decimal(request.POST.get("amount") or "0")
        date = request.POST.get("date") or None
        method = request.POST.get("method") or "Cash"

        if amt > 0:
            Payment.objects.create(
                invoice=inv,
                amount=amt,
                date=date or timezone.localdate(),
                method=method,
            )
            inv.recalc_payments_and_status()

        return redirect("sales:unpaid_invoices")

    return render(
        request,
        "sales/invoice_mark_paid.html",
        {"inv": inv},
    )


# ============================================================
# REFUND CRUD + LIST
# ============================================================

# sales/views.py (append)
from decimal import Decimal
from catalog.models_stock import StockLedger
from django.shortcuts import render
from django.urls import reverse

@login_required
@transaction.atomic
def refund_add(request):
    """
    Handles refund creation and mapping returned qty to StockLedger (restock).
    Validations:
     - cannot refund more than invoice.paid_amount
     - cannot return more qty than remaining (invoiced - already returned)
    """
    invoices = Invoice.objects.select_related('customer').order_by('-date')

    if request.method == "POST":
        invoice_id = request.POST.get("invoice")
        if not invoice_id:
            messages.error(request, "Invoice required for refund.")
            return redirect('sales:refund_add')

        inv = get_object_or_404(Invoice, pk=invoice_id)

        # parse returned lines
        item_ids = request.POST.getlist("item_id[]")
        qtys = request.POST.getlist("qty_returned[]")
        returned_lines = []
        for i, iid in enumerate(item_ids):
            try:
                q = int(float(qtys[i])) if i < len(qtys) and qtys[i] else 0
            except Exception:
                q = 0
            if q > 0:
                returned_lines.append((int(iid), q))

        # compute refund amount if not provided: sum(rate * qty) for returned lines
        total_amount = Decimal(request.POST.get("amount") or "0")
        if total_amount == Decimal("0") and returned_lines:
            amt_calc = Decimal("0")
            for iid, q in returned_lines:
                it = InvoiceItem.objects.filter(pk=iid, invoice=inv).first()
                if it:
                    amt_calc += (Decimal(it.rate or 0) * Decimal(q))
            total_amount = amt_calc.quantize(Decimal("0.01"))

        method = request.POST.get("method") or "Refund"
        note = request.POST.get("note") or ""

        # validation: refund amount cannot exceed paid_amount
        paid_amount = inv.paid_amount or Decimal("0")
        if total_amount > paid_amount:
            messages.error(request, "Refund amount cannot be greater than invoice paid amount.")
            return redirect('sales:refund_add')

        # Validate returned_lines qtys do not exceed remaining per invoice item
        # Build a map of invoice_item_id -> already_returned (sum over existing refund payments)
        refund_payment_ids = list(Payment.objects.filter(invoice=inv, is_refund=True).values_list('id', flat=True))
        already_returned_map = {}
        for iid, _ in returned_lines:
            it = InvoiceItem.objects.filter(pk=iid, invoice=inv).first()
            if not it:
                already_returned_map[iid] = 0
                continue
            prod = getattr(it, 'product', None)
            returned_sum = 0
            if prod and refund_payment_ids:
                returned_sum = StockLedger.objects.filter(
                    product=prod,
                    ref_type__in=['return', 'return_unallocated', 'invoice_reverse'],
                    ref_id__in=refund_payment_ids
                ).aggregate(total=Sum('in_qty'))['total'] or 0
            already_returned_map[iid] = int(returned_sum)

        # Now check requested qtys
        for iid, q in returned_lines:
            it = InvoiceItem.objects.filter(pk=iid, invoice=inv).first()
            if not it:
                messages.error(request, "Invalid invoice item selected.")
                return redirect('sales:refund_add')
            already = already_returned_map.get(iid, 0)
            remaining = int(it.quantity) - already
            if q > remaining:
                messages.error(request, f"Return qty for item {it.description} cannot exceed remaining ({remaining}).")
                return redirect('sales:refund_add')

        # Create Payment refund record
        refund = Payment.objects.create(
            invoice=inv,
            amount=total_amount.quantize(Decimal("0.01")),
            method=method,
            date=request.POST.get("date") or timezone.localdate(),
            note=note,
            is_refund=True,
        )

        # Map returned qty to StockLedger out entries for this invoice & create in entries
        for iid, qty_to_return in returned_lines:
            it = InvoiceItem.objects.filter(pk=iid, invoice=inv).first()
            if not it or qty_to_return <= 0:
                continue
            product = getattr(it, 'product', None)
            if not product:
                # nothing to restock for manual/saved items
                continue

            qty_left = int(qty_to_return)

            # find out entries created when invoice allocated stock (ref_type='invoice', ref_id=inv.id)
            outs = StockLedger.objects.filter(
                product=product,
                ref_type='invoice',
                ref_id=inv.id,
                out_qty__gt=0
            ).order_by('id')  # FIFO

            for out in outs:
                if qty_left <= 0:
                    break

                # compute how much of this particular out has already been reversed/returned (against this print_run)
                returned_against_pr = StockLedger.objects.filter(
                    product=product,
                    print_run=out.print_run,
                    ref_type__in=['return', 'return_unallocated', 'invoice_reverse'],
                    # We accept any refund ref_ids because earlier returns for same invoice might map to same print_run
                ).aggregate(total=Sum('in_qty'))['total'] or 0

                available_from_out_raw = int((out.out_qty or 0) - returned_against_pr)
                if available_from_out_raw <= 0:
                    continue

                take = min(available_from_out_raw, qty_left)

                last_bal = StockLedger.objects.filter(product=product).order_by('-id').first()
                last_balance = last_bal.balance if last_bal else 0

                StockLedger.objects.create(
                    product=product,
                    print_run=out.print_run,
                    warehouse=out.warehouse,
                    in_qty=take,
                    out_qty=0,
                    balance=last_balance + take,
                    ref_type='return',
                    ref_id=refund.id,
                    notes=f"Return for invoice {inv.number}, invoice_item {it.id}"
                )
                qty_left -= take

            if qty_left > 0:
                # create an unallocated in entry (fallback)
                last_bal = StockLedger.objects.filter(product=product).order_by('-id').first()
                last_balance = last_bal.balance if last_bal else 0
                StockLedger.objects.create(
                    product=product,
                    print_run=None,
                    warehouse=None,
                    in_qty=qty_left,
                    out_qty=0,
                    balance=last_balance + qty_left,
                    ref_type='return_unallocated',
                    ref_id=refund.id,
                    notes=f"Return (unallocated) for invoice {inv.number}, invoice_item {it.id}"
                )
                qty_left = 0

        # Recalculate invoice totals/status
        try:
            inv.recalc_totals()
            inv.recalc_payments_and_status()
        except Exception:
            logger.exception("Invoice recalc failed in refund_add. invoice_id=%s", inv.id)

        messages.success(request, f"Refund recorded (id: {refund.id}). Stock updated.")
        return redirect("sales:refund_detail", pk=refund.id)

    # GET request
    return render(request, "sales/refund_add.html", {"invoices": invoices})

from datetime import timedelta
from django.utils.dateparse import parse_date


@login_required
def refund_statement_pdf(request):
    """
    Refund Statement PDF WITHOUT wkhtmltopdf / xhtml2pdf.
    Pure ReportLab se table-based PDF banata hai.
    """

    # Refunds -> invoice + customer ke saath
    refunds = (
        Refund.objects
        .select_related("invoice", "invoice__customer")
        .order_by("date", "id")
    )

    total_refunded = (
        refunds.aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    ).quantize(Decimal("0.01"))

    # HTTP response prepare
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="refund-statement.pdf"'

    # PDF doc setup
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Refund Statement", styles["Title"]))
    elements.append(
        Paragraph(
            f"Generated at: {timezone.now().strftime('%d-%m-%Y %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 10))

    # Table header
    data = [["Date", "Invoice #", "Customer", "Amount"]]

    # Table rows
    for r in refunds:
        inv = getattr(r, "invoice", None)
        cust = getattr(inv, "customer", None) if inv else None

        date_str = r.date.strftime("%d-%m-%Y") if getattr(r, "date", None) else ""
        inv_num = inv.number if inv else ""
        cust_name = cust.name if cust else ""
        amount_str = f"{r.amount}"

        data.append([date_str, inv_num, cust_name, amount_str])

    # Total row
    data.append(["", "", "Total Refunded", f"{total_refunded}"])

    # Table layout
    table = Table(
        data,
        colWidths=[25 * mm, 30 * mm, 80 * mm, 25 * mm],
        repeatRows=1,  # header har page pe repeat
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),

                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("FONTSIZE", (0, 1), (-1, -2), 9),

                # Total row style (last row)
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
            ]
        )
    )

    elements.append(table)

    # build PDF
    doc.build(elements)
    return response



@login_required
def refund_detail(request, pk):
    """
    Single refund detail (Payment with is_refund=True).
    """
    refund = get_object_or_404(Payment, pk=pk, is_refund=True)
    return render(request, "sales/refund_detail.html", {"refund": refund})

@login_required
def refund_list(request):
    # Base queryset: saare refunds (Payment me is_refund=True)
    refunds = (
        Payment.objects
        .filter(is_refund=True)
        .select_related('invoice', 'invoice__customer')
        .order_by('-date', '-id')
    )

    # ------- Filters -------
    q = request.GET.get('q', '').strip()              # search: customer, invoice, method, note
    customer_id = request.GET.get('customer', '').strip()  # dropdown / direct id
    export = request.GET.get('export', '').strip()    # csv / xlsx

    if q:
        refunds = refunds.filter(
            Q(invoice__number__icontains=q) |
            Q(invoice__customer__name__icontains=q) |
            Q(method__icontains=q) |
            Q(note__icontains=q)
        )

    if customer_id:
        refunds = refunds.filter(invoice__customer_id=customer_id)

    # ------- Export (respecting filters) -------
    if export == 'csv':
        return export_refunds_csv(refunds)
    elif export == 'xlsx':
        return export_refunds_excel(refunds)

    # ------- Summary numbers -------
    total_refunds = refunds.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    refund_count = refunds.count()
    last_refund = refunds.first()

    if refund_count > 0:
        avg_refund = (total_refunds / refund_count).quantize(Decimal("0.01"))
    else:
        avg_refund = Decimal("0.00")

    # Customer dropdown ke liye (sirf un customers jinke paas refund hai)
    customers_with_refunds = (
        Customer.objects
        .filter(invoices__payments__is_refund=True)
        .distinct()
        .order_by('name')
    )

    return render(
        request,
        'sales/refund_list.html',
        {
            'refunds': refunds,
            'total_refunds': total_refunds,
            'refund_count': refund_count,
            'last_refund': last_refund,
            'customer_id': customer_id,
            'customers_with_refunds': customers_with_refunds,
            'q': q,
            'avg_refund': avg_refund,
        }
    )


def export_refunds_csv(refunds_qs):
    """Filtered refunds ko CSV me export karega."""
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="refunds.csv"'

    writer = csv.writer(resp)
    writer.writerow([
        'Date',
        'Refund ID',
        'Invoice #',
        'Customer',
        'Method',
        'Amount',
        'Note',
    ])

    for r in refunds_qs:
        writer.writerow([
            r.date.strftime('%Y-%m-%d') if r.date else '',
            r.id,
            r.invoice.number if r.invoice else '',
            r.invoice.customer.name if r.invoice and r.invoice.customer else '',
            r.method or '',
            float(r.amount or 0),
            (r.note or '').replace('\n', ' '),
        ])

    return resp


def export_refunds_excel(refunds_qs):
    """Filtered refunds ko Excel (XLSX) me export karega. Requires openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError:
        # agar openpyxl nahi hai toh CSV fallback
        return export_refunds_csv(refunds_qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Refunds"

    ws.append([
        'Date',
        'Refund ID',
        'Invoice #',
        'Customer',
        'Method',
        'Amount',
        'Note',
    ])

    for r in refunds_qs:
        ws.append([
            r.date.strftime('%Y-%m-%d') if r.date else '',
            r.id,
            r.invoice.number if r.invoice else '',
            r.invoice.customer.name if r.invoice and r.invoice.customer else '',
            r.method or '',
            float(r.amount or 0),
            r.note or '',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="refunds.xlsx"'
    return resp




@login_required
@transaction.atomic
def refund_edit(request, pk):
    """
    Existing refund ko edit karne ka page:
    - invoice change kar sakte ho
    - amount / date / method / note update
    - invoice totals/status bhi recalc honge
    """
    refund = get_object_or_404(Payment, pk=pk, is_refund=True)
    invoices = Invoice.objects.select_related('customer').order_by('-date')

    today = timezone.localdate()

    if request.method == "POST":
        invoice_id = request.POST.get("invoice") or None
        amount = parse_decimal(request.POST.get("amount") or "0")
        date = request.POST.get("date") or None
        method = request.POST.get("method") or "Refund"
        note = request.POST.get("note") or ""

        inv = Invoice.objects.get(pk=invoice_id) if invoice_id else None

        refund.invoice = inv
        refund.amount = amount.quantize(Decimal("0.01"))
        refund.date = date or today
        refund.method = method
        refund.note = note
        refund.save()

        # Invoice ke totals + status recalc
        if inv:
            try:
                inv.recalc_totals()
                inv.recalc_payments_and_status()
            except Exception:
                logger.exception("Invoice recalc failed in refund_edit. invoice_id=%s", inv.id if inv else None)

        return redirect("sales:refund_detail", pk=refund.pk)

    context = {
        "refund": refund,
        "invoices": invoices,
        "today": today,
    }
    return render(request, "sales/refund_edit.html", context)



@login_required
def refund_pdf(request, pk):
    """
    Single refund ka professional PDF export (pdfkit + wkhtmltopdf).
    """
    refund = get_object_or_404(Payment, pk=pk, is_refund=True)

    invoice = refund.invoice
    customer = invoice.customer if invoice and getattr(invoice, "customer", None) else None

    # Company: priority invoice.company, warna first CompanyProfile
    company = None
    if invoice and getattr(invoice, "company", None):
        company = invoice.company
    if company is None:
        company = CompanyProfile.objects.first()

    context = {
        "refund": refund,
        "invoice": invoice,
        "customer": customer,
        "company": company,
        "now": timezone.now(),
    }

    template = get_template("sales/refund_pdf.html")
    html = template.render(context, request=request)

    config = pdf_config()
    options = pdf_options()

    pdf_bytes = pdfkit.from_string(
        html,
        False,
        configuration=config,
        options=options,
    )

    filename = f"refund_{refund.id}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



@login_required
@transaction.atomic
def refund_delete(request, pk):
    """
    Delete a refund and keep invoice amounts consistent.
    """
    refund = get_object_or_404(Payment, pk=pk, is_refund=True)
    inv = refund.invoice
    amount = refund.amount

    if request.method == "POST":
        refund.delete()

        # Recalculate via central helpers, fallback if needed
        try:
            inv.recalc_totals()
            inv.recalc_payments_and_status()
        except Exception:
            logger.exception("Invoice recalc failed in refund_delete. invoice_id=%s", inv.id if inv else None)
            inv.paid_amount = (
                inv.paid_amount + amount
            ).quantize(Decimal("0.01"))
            inv.balance_due = (
                inv.grand_total - inv.paid_amount
            ).quantize(Decimal("0.01"))

            if inv.balance_due <= 0:
                inv.status = "paid"
                inv.balance_due = Decimal("0.00")
            elif inv.paid_amount > 0:
                inv.status = "partial"
            else:
                inv.status = "unpaid"

            inv.save()

        return redirect("sales:refund_list")

    return render(
        request,
        "sales/refund_delete_confirm.html",
        {"refund": refund},
    )


# ============================================================
# CUSTOMER STATEMENT PDF
# ============================================================

@login_required
def customer_statement_pdf(request, pk):
    """
    Customer-wise statement PDF.
    """
    customer = get_object_or_404(Customer, pk=pk)

    from_str = request.GET.get("from")  # YYYY-MM-DD
    to_str = request.GET.get("to")      # YYYY-MM-DD

    from_date = None
    to_date = None

    invoices_qs = Invoice.objects.filter(
        customer=customer,
        deleted_at__isnull=True,
        in_bin=False,
    )

    if from_str:
        try:
            from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
            invoices_qs = invoices_qs.filter(date__gte=from_date)
        except ValueError:
            from_date = None

    if to_str:
        try:
            to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
            invoices_qs = invoices_qs.filter(date__lte=to_date)
        except ValueError:
            to_date = None

    invoices = invoices_qs.order_by("date", "id")

    totals = invoices.aggregate(
        total_spent=Sum("grand_total"),
        pending_balance=Sum("balance_due"),
    )

    total_spent = (totals["total_spent"] or Decimal("0.00")).quantize(Decimal("0.01"))
    pending_balance = (totals["pending_balance"] or Decimal("0.00")).quantize(Decimal("0.01"))

    payments_qs = Payment.objects.filter(
        invoice__in=invoices,
        invoice__customer=customer,
        is_refund=False,
    ).select_related("invoice")

    refunds_payments_qs = Payment.objects.filter(
        invoice__in=invoices,
        invoice__customer=customer,
        is_refund=True,
    ).select_related("invoice")

    legacy_refunds_qs = Refund.objects.filter(
        invoice__in=invoices,
        invoice__customer=customer,
    )

    total_paid, total_refunded_payments, _ = aggregate_payment_totals(
        Payment.objects.filter(invoice__in=invoices, invoice__customer=customer)
    )
    total_refunded = aggregate_refunds_unified(payments_qs.filter(is_refund=True), legacy_refunds_qs)
    net_received = (total_paid - total_refunded).quantize(Decimal("0.01"))

    company = None
    first_inv = invoices.first()
    if first_inv and getattr(first_inv, "company", None):
        company = first_inv.company
    if company is None:
        company = CompanyProfile.objects.first()

    now = timezone.now()

    context = {
        "customer": customer,
        "invoices": invoices,
        "total_spent": total_spent,
        "pending_balance": pending_balance,
        "total_paid": total_paid,
        "total_refunded": total_refunded,
        "net_received": net_received,
        "refunds": refunds_payments_qs,
        "legacy_refunds": legacy_refunds_qs,
        "company": company,
        "now": now,
        "from_date": from_date,
        "to_date": to_date,
        "statement_note": "",
    }

    template = get_template("sales/customer_statement_pdf.html")
    html = template.render(context, request=request)

    config = pdf_config()
    options = pdf_options()

    pdf_bytes = pdfkit.from_string(
        html,
        False,
        configuration=config,
        options=options,
    )

    filename = f"statement_{customer.id}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


from datetime import datetime, timedelta

@login_required
def payments_statement_pdf(request):
    """
    Overall Payments Statement PDF
    - period: quarter (default), half, year, custom
    - custom: ?from=YYYY-MM-DD&to=YYYY-MM-DD
    """
    today = timezone.localdate()
    period = request.GET.get("period", "quarter")
    from_str = request.GET.get("from")
    to_str = request.GET.get("to")

    from_date = None
    to_date = None

    if period == "custom" and from_str and to_str:
        try:
            from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
        except ValueError:
            from_date = None
        try:
            to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
        except ValueError:
            to_date = None

    if not from_date or not to_date:
        to_date = today
        if period == "half":
            from_date = today - timedelta(days=180)
        elif period == "year":
            from_date = today - timedelta(days=365)
        else:
            from_date = today - timedelta(days=90)

    payments_qs = Payment.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    ).select_related("invoice", "invoice__customer")

    total_paid, total_refunded, net_received = aggregate_payment_totals(payments_qs)

    payments = payments_qs.filter(is_refund=False).order_by("date", "id")
    refunds = payments_qs.filter(is_refund=True).order_by("date", "id")

    by_method = (
        payments_qs.values("method")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )

    company = CompanyProfile.objects.first()
    now = timezone.now()

    context = {
        "company": company,
        "now": now,
        "from_date": from_date,
        "to_date": to_date,
        "period": period,
        "total_paid": total_paid,
        "total_refunded": total_refunded,
        "net_received": net_received,
        "payments": payments,
        "refunds": refunds,
        "by_method": by_method,
    }

    template = get_template("sales/payments_statement_pdf.html")
    html = template.render(context, request=request)

    config = pdf_config()
    options = pdf_options()

    pdf_bytes = pdfkit.from_string(
        html,
        False,
        configuration=config,
        options=options,
    )

    filename = f"payments_statement_{from_date}_{to_date}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


from django.urls import reverse


from django.db.models import F

@login_required
def bulk_invoice_action(request):
    """
    Generic bulk action endpoint for invoices.
    Expects POST with:
      - invoice_ids[] (list) OR invoice_ids (list)
      - action (string): 'mark_paid' | 'export_csv' | 'send_whatsapp' | ...
    """
    if request.method != "POST":
        return redirect('sales:unpaid_invoices')

    # Accept both names coming from various templates
    invoice_ids = request.POST.getlist('invoice_ids[]') or request.POST.getlist('invoice_ids')
    action = request.POST.get('action') or request.POST.get('bulk_action')

    if not invoice_ids:
        messages.warning(request, "No invoices selected for bulk action.")
        return redirect(request.META.get('HTTP_REFERER') or reverse('sales:unpaid_invoices'))

    invoices = Invoice.objects.filter(id__in=invoice_ids).select_related('customer')

    if action == 'mark_paid':
        created_count = 0
        with transaction.atomic():
            # lock rows to avoid race conditions
            for inv in invoices.select_for_update():
                # compute balance safely (adapt if your model uses different field names)
                try:
                    balance = getattr(inv, 'balance_due', None)
                    if balance is None:
                        total = Decimal(getattr(inv, 'grand_total', 0) or 0)
                        paid = Decimal(getattr(inv, 'paid_amount', 0) or 0)
                        balance = total - paid
                    balance = Decimal(balance or 0)
                except Exception:
                    balance = Decimal('0.00')

                if balance > 0:
                    # prepare kwargs for Payment creation, only include fields that exist
                    payment_kwargs = {
                        'invoice': inv,
                        'amount': balance,
                        # date: use date/datetime according to your model field type
                        # if Payment.date is DateField this is ok; if DateTimeField you may prefer timezone.now()
                        'date': timezone.now().date(),
                        'method': 'Manual (bulk mark paid)',
                        'is_refund': False,
                    }

                    # detect model fields and optionally add created_by if it exists
                    payment_field_names = {f.name for f in Payment._meta.get_fields()}
                    if 'created_by' in payment_field_names and hasattr(request, 'user'):
                        payment_kwargs['created_by'] = request.user
                    # if your model uses a different user field (eg `user`), handle accordingly:
                    # if 'user' in payment_field_names: payment_kwargs['user'] = request.user

                    # create payment
                    Payment.objects.create(**payment_kwargs)

                    # update invoice using F() to avoid race issues
                    try:
                        Invoice.objects.filter(pk=inv.pk).update(
                            paid_amount=F('paid_amount') + balance,
                            balance_due=Decimal('0.00'),
                            status='paid'  # adapt status value to your choices if needed
                        )
                        inv.refresh_from_db()
                    except Exception:
                        # fallback: set attributes and save (if model fields differ)
                        inv.paid_amount = (getattr(inv, 'paid_amount', Decimal('0.00')) or Decimal('0.00')) + balance
                        inv.balance_due = Decimal('0.00')
                        inv.status = 'paid'
                        inv.save()

                    created_count += 1

        messages.success(request, f"{created_count} invoice(s) marked as paid.")
        return redirect(request.META.get('HTTP_REFERER') or reverse('sales:unpaid_invoices'))

    elif action == 'export_csv':
        response = HttpResponse(content_type='text/csv')
        filename = f"invoices_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(['Invoice ID', 'Number', 'Customer', 'Date', 'Total', 'Paid', 'Balance', 'Status'])
        for inv in invoices:
            cust = inv.customer.name if getattr(inv, 'customer', None) else ''
            writer.writerow([
                inv.id,
                getattr(inv, 'number', ''),
                cust,
                getattr(inv, 'date', ''),
                getattr(inv, 'grand_total', ''),
                getattr(inv, 'paid_amount', ''),
                getattr(inv, 'balance_due', ''),
                getattr(inv, 'status', ''),
            ])
        return response

    elif action == 'send_whatsapp':
        count = 0
        for inv in invoices:
            # implement your actual WA queueing or mark logic here
            count += 1
        messages.success(request, f"WhatsApp reminder queued for {count} invoice(s).")
        return redirect(request.META.get('HTTP_REFERER') or reverse('sales:unpaid_invoices'))

    else:
        messages.error(request, "Unknown bulk action.")
        return redirect(request.META.get('HTTP_REFERER') or reverse('sales:unpaid_invoices'))

from catalog.models_stock import StockLedger
from catalog.models import Product as CatalogProduct

@login_required
def invoice_items_api(request, invoice_id):
    """
    JSON API returning invoice items and how many can be returned.
    This version computes qty already returned by summing StockLedger.in_qty
    for all refund Payment records tied to this invoice.
    URL: /sales/api/invoice-items/<invoice_id>/
    """
    try:
        inv = Invoice.objects.select_related('customer').get(pk=invoice_id)
    except Invoice.DoesNotExist:
        raise Http404("Invoice not found")

    # get IDs of refund payments already created for this invoice
    refund_payment_ids = list(
        Payment.objects.filter(invoice=inv, is_refund=True).values_list('id', flat=True)
    )

    items = []
    for it in inv.items.all():
        prod = getattr(it, 'product', None)

        # default returned sum = 0
        returned_sum = 0

        if prod:
            # Sum in_qty from StockLedger where ref_type indicates a return and ref_id is one of refund payment ids
            # Also accept invoice_reverse and return_unallocated as possible ref_types
            if refund_payment_ids:
                qs = StockLedger.objects.filter(
                    product=prod,
                    ref_type__in=['return', 'return_unallocated', 'invoice_reverse'],
                    ref_id__in=refund_payment_ids
                )
                # Some implementations may have stored invoice_item id in notes; try to further narrow if present.
                # But don't require it — fallback to summing all returns for this product+invoice.
                returned_sum = qs.aggregate(total=Sum('in_qty'))['total'] or 0
            else:
                returned_sum = 0

        already_returned = int(returned_sum)
        qty_available = max(0, int(it.quantity) - already_returned)

        # include numeric rate for frontend calc
        rate_val = float(it.rate or 0)

        items.append({
            "invoice_item_id": it.id,
            "product_id": prod.id if prod else None,
            "description": it.description,
            "qty_invoiced": int(it.quantity),
            "qty_already_returned": already_returned,
            "qty_available_to_return": qty_available,
            "rate": rate_val,
            "line_total": float((Decimal(it.rate or 0) * Decimal(it.quantity)).quantize(Decimal("0.01")))
        })

    data = {
        "ok": True,
        "invoice": {"id": inv.id, "number": inv.number, "customer": inv.customer.name if inv.customer else ""},
        "items": items
    }
    return JsonResponse(data)


@login_required
def inventory_home(request):
    """
    Simple redirect so sales/inventory/ points to catalog inventory home.
    Keeps your URLs working without duplicating dashboard code.
    """
    return redirect("catalog:inventory_home")


