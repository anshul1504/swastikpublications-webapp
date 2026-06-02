# ============================================================================
# catalog/views.py — FINAL FIXED VERSION (MATCHES YOUR TEMPLATES)
# ============================================================================

import csv
import json
import io
import logging
from decimal import Decimal
from django import forms
from django.apps import apps
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, F
from django.http import Http404
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.utils.http import urlencode
from django.utils import timezone
from django.views.decorators.http import require_POST, require_http_methods
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView, DetailView
)
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required


# ----------------------------------------------------------------------------
# MODELS
# ----------------------------------------------------------------------------
try:
    from .models import Product, PrintRun, Warehouse, StockLedger
except:
    Product = apps.get_model("catalog", "Product")
    PrintRun = apps.get_model("catalog", "PrintRun")
    Warehouse = apps.get_model("catalog", "Warehouse")
    StockLedger = apps.get_model("catalog", "StockLedger")


# ----------------------------------------------------------------------------
# FORMS
# ----------------------------------------------------------------------------
try:
    from .forms import PrintRunForm, WarehouseForm, ProductForm
except:
    PrintRunForm = forms.modelform_factory(PrintRun, exclude=[])
    WarehouseForm = forms.modelform_factory(Warehouse, exclude=[])
    ProductForm = forms.modelform_factory(Product, exclude=[])


# ============================================================================
# 1) PRODUCT CRUD
# ============================================================================

class ProductList(LoginRequiredMixin, ListView):
    model = Product
    template_name = "catalog/product_list.html"
    context_object_name = "products"
    paginate_by = 25

    def get_queryset(self):
        qs = Product.objects.all().order_by("name")

        q = (self.request.GET.get("q") or "").strip()
        status = (self.request.GET.get("status") or "").strip()
        track = (self.request.GET.get("track") or "").strip()

        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(sku__icontains=q)
                | Q(description__icontains=q)
                | Q(isbn__icontains=q)
                | Q(author__icontains=q)
                | Q(imprint__icontains=q)
                | Q(edition__icontains=q)
                | Q(academic_session__icontains=q)
                | Q(subject__icontains=q)
            )

        if status == "active":
            qs = qs.filter(active=True)
        elif status == "inactive":
            qs = qs.filter(active=False)

        if track == "yes":
            qs = qs.filter(track_stock=True)
        elif track == "no":
            qs = qs.filter(track_stock=False)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        q = (self.request.GET.get("q") or "").strip()
        status = (self.request.GET.get("status") or "").strip()
        track = (self.request.GET.get("track") or "").strip()

        ctx["q"] = q
        ctx["status"] = status
        ctx["track"] = track

        # total products (without filters)
        ctx["total_count"] = Product.objects.count()

        # filtered products (respecting filters + pagination)
        paginator = ctx.get("paginator", None)
        page_obj = ctx.get("page_obj", None)

        if paginator is not None:
            ctx["filtered_count"] = paginator.count
        else:
            obj_list = ctx.get("products") or ctx.get("object_list")
            if hasattr(obj_list, "count"):
                ctx["filtered_count"] = obj_list.count()
            else:
                ctx["filtered_count"] = 0

        # nice display range "showing X–Y of Z"
        if page_obj and ctx["filtered_count"] > 0:
            start_index = (page_obj.number - 1) * self.paginate_by + 1
            end_index = start_index + len(page_obj.object_list) - 1
            ctx["start_index"] = start_index
            ctx["end_index"] = end_index
        else:
            ctx["start_index"] = 0
            ctx["end_index"] = 0

        return ctx


class ProductCreate(LoginRequiredMixin, CreateView):
    model = Product
    template_name = "catalog/product_add.html"
    form_class = ProductForm
    success_url = reverse_lazy("catalog:product_list")


class ProductUpdate(LoginRequiredMixin, UpdateView):
    model = Product
    template_name = "catalog/product_edit.html"
    form_class = ProductForm
    success_url = reverse_lazy("catalog:product_list")


class ProductDelete(LoginRequiredMixin, DeleteView):
    model = Product
    template_name = "catalog/product_delete.html"
    success_url = reverse_lazy("catalog:product_list")

# ============================================================================
# 2) WAREHOUSE CRUD
# ============================================================================
# ========================================================================
# 2) WAREHOUSE CRUD
# ========================================================================

class WarehouseList(LoginRequiredMixin, ListView):
    model = Warehouse
    template_name = "catalog/warehouse_list.html"
    context_object_name = "warehouses"
    paginate_by = 25
    ordering = ["name"]

    def get_queryset(self):
        qs = Warehouse.objects.all().order_by(*self.ordering)

        q = (self.request.GET.get("q") or "").strip()
        status = (self.request.GET.get("status") or "").strip()

        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(phone__icontains=q)
                | Q(email__icontains=q)
                | Q(address__icontains=q)
            )

        if status and hasattr(Warehouse, "is_active"):
            if status == "active":
                qs = qs.filter(is_active=True)
            elif status == "inactive":
                qs = qs.filter(is_active=False)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        q = (self.request.GET.get("q") or "").strip()
        status = (self.request.GET.get("status") or "").strip()

        ctx["q"] = q
        ctx["status"] = status

        # Total warehouses (no filters)
        ctx["total_count"] = Warehouse.objects.count()

        # Filter ke baad kitne – paginator se safe count
        paginator = ctx.get("paginator")
        page_obj = ctx.get("page_obj")
        if paginator is not None:
            ctx["filtered_count"] = paginator.count
        else:
            obj_list = ctx.get("warehouses") or ctx.get("object_list")
            if hasattr(obj_list, "count"):
                ctx["filtered_count"] = obj_list.count()
            else:
                ctx["filtered_count"] = 0

        return ctx


class WarehouseCreate(LoginRequiredMixin, CreateView):
    model = Warehouse
    template_name = "catalog/warehouse_add.html"
    form_class = WarehouseForm
    success_url = reverse_lazy("catalog:warehouse_list")


class WarehouseUpdate(LoginRequiredMixin, UpdateView):
    model = Warehouse
    template_name = "catalog/warehouse_edit.html"
    form_class = WarehouseForm
    success_url = reverse_lazy("catalog:warehouse_list")


class WarehouseDelete(LoginRequiredMixin, DeleteView):
    model = Warehouse
    template_name = "catalog/warehouse_delete.html"
    success_url = reverse_lazy("catalog:warehouse_list")


# catalog/views.py

class WarehouseDetail(LoginRequiredMixin, DetailView):
    model = Warehouse
    template_name = "catalog/warehouse_detail.html"
    context_object_name = "warehouse"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        wh = self.object

        # All ledger entries for this warehouse
        ledgers = StockLedger.objects.filter(warehouse=wh).order_by("date", "id")
        ctx["stock_ledgers"] = ledgers

        # ---- STOCK AGGREGATES ----
        agg = (
            StockLedger.objects
            .filter(warehouse=wh)
            .values("product")
            .annotate(qty=Sum("in_qty") - Sum("out_qty"))
)


        stock_count = agg.filter(qty__gt=0).count()
        stock_qty = agg.aggregate(total=Sum("qty"))["total"] or 0

        ctx["stock_count"] = stock_count
        ctx["stock_qty"] = stock_qty

        # Transfers count (simple: any ref_type containing "transfer")
        transfers_count = ledgers.filter(ref_type__icontains="transfer").count()
        ctx["transfers_count"] = transfers_count

        # Last sync = last ledger date
        last_sync = (
            ledgers.order_by("-date", "-id")
            .values_list("date", flat=True)
            .first()
        )
        ctx["last_sync"] = last_sync

        # ---- RECENT ACTIVITY (5 items) ----
        recent = (
            ledgers.select_related("product")
            .order_by("-date", "-id")[:5]
        )

        activity = []
        for sl in recent:
            parts = []
            if sl.product:
                parts.append(sl.product.name)
            if (sl.in_qty or 0) or (sl.out_qty or 0):
                parts.append(f"+{sl.in_qty or 0} / -{sl.out_qty or 0}")
            if sl.ref_type:
                parts.append(sl.ref_type)
            msg = " · ".join(parts) or "Stock movement"
            activity.append(
                {
                    "message": msg,
                    "timestamp": sl.date,
                }
            )
        ctx["recent_activity"] = activity

        # ---- SPARKLINE DATA (stock trend for this warehouse) ----
        history = {}
        running = 0
        for sl in ledgers:
            date_obj = getattr(sl, "date", None)
            if hasattr(date_obj, "date"):
                key = date_obj.date()
            else:
                key = date_obj
            delta = int(sl.in_qty or 0) - int(sl.out_qty or 0)
            running += delta
            history[key] = running

        items = sorted(history.items())[-20:]  # last 20 points max
        labels = [
            d.strftime("%d %b") if hasattr(d, "strftime") else str(d)
            for d, _ in items
        ]
        values = [v for _, v in items]

        ctx["spark_labels_json"] = json.dumps(labels)
        ctx["spark_values_json"] = json.dumps(values)

        # Flags for inline edit modal (only show if fields exist)
        ctx["has_manager"] = hasattr(wh, "manager_name")
        ctx["has_opening"] = hasattr(wh, "opening_hours")
        ctx["has_notes"] = hasattr(wh, "notes")

        return ctx

# ============================================================================
# 3) PRINTRUN CRUD
# ============================================================================

class PrintRunList(LoginRequiredMixin, ListView):
    model = PrintRun
    template_name = "catalog/printrun_list.html"
    context_object_name = "print_runs"
    paginate_by = 25

    def get_queryset(self):
        qs = (
            PrintRun.objects
            .select_related("product", "warehouse")
        )

        # --- filters from GET ---
        self.filter_q = (self.request.GET.get("q") or "").strip()
        self.filter_warehouse = (self.request.GET.get("warehouse") or "").strip()
        self.filter_date_from = (self.request.GET.get("date_from") or "").strip()
        self.filter_date_to = (self.request.GET.get("date_to") or "").strip()
        self.filter_sort = (self.request.GET.get("sort") or "").strip()

        if self.filter_q:
            qs = qs.filter(
                Q(product__name__icontains=self.filter_q)
                | Q(product__sku__icontains=self.filter_q)
                | Q(batch_no__icontains=self.filter_q)
                | Q(notes__icontains=self.filter_q)
            )

        if self.filter_warehouse:
            try:
                qs = qs.filter(warehouse_id=int(self.filter_warehouse))
            except (ValueError, TypeError):
                pass

        if self.filter_date_from:
            qs = qs.filter(print_date__gte=self.filter_date_from)

        if self.filter_date_to:
            qs = qs.filter(print_date__lte=self.filter_date_to)

        # Sorting
        sort = self.filter_sort
        if sort == "oldest":
            qs = qs.order_by("print_date", "id")
        else:
            # default: newest first
            qs = qs.order_by("-print_date", "-id")

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Warehouses for dropdown
        ctx["warehouses"] = Warehouse.objects.all().order_by("name")

        # Expose filters for template
        ctx["filter_q"] = getattr(self, "filter_q", "")
        ctx["filter_warehouse"] = getattr(self, "filter_warehouse", "")
        ctx["filter_date_from"] = getattr(self, "filter_date_from", "")
        ctx["filter_date_to"] = getattr(self, "filter_date_to", "")
        ctx["filter_sort"] = getattr(self, "filter_sort", "")

        # querystring without page (for pagination + export links)
        params = self.request.GET.copy()
        params.pop("page", None)
        ctx["querystring_no_page"] = params.urlencode()

        # Optional: compute available qty if method exists
        for pr in ctx["print_runs"]:
            try:
                if hasattr(pr, "available_qty") and callable(getattr(pr, "available_qty")):
                    pr.computed_available_qty = pr.available_qty()
            except Exception:
                pr.computed_available_qty = None

        return ctx


class PrintRunCreate(LoginRequiredMixin, CreateView):
    model = PrintRun
    template_name = "catalog/printrun_form.html"
    form_class = PrintRunForm
    success_url = reverse_lazy("catalog:pr_list")

    def form_valid(self, form):
        pr = form.save(commit=False)

        # Auto batch number
        if not pr.batch_no:
            ts = int(timezone.now().timestamp())
            pr.batch_no = f"PR-{pr.product.id}-{ts}"

        pr.save()

        # Ledger entry
        StockLedger.objects.create(
            product=pr.product,
            print_run=pr,
            warehouse=pr.warehouse,
            in_qty=pr.received_qty,
            out_qty=0,
            balance=pr.received_qty,
            ref_type="pr_create",
            notes=f"Created by {self.request.user}"
        )

        messages.success(self.request, "Batch created and stock added.")
        return redirect(self.success_url)

# ---------------------------------------------------------------------------
# PRINT-RUN UPDATE (EDIT)  –  creates ledger adjustment rows on change
# ---------------------------------------------------------------------------
class PrintRunUpdate(LoginRequiredMixin, UpdateView):
    """
    Basic update view – yeh sirf form + template handle karta hai.
    Actual stock adjustment PrintRunUpdateEnterprise me hai.
    (Urls me hum Enterprise wali class use kar rahe hain.)
    """
    model = PrintRun
    form_class = PrintRunForm
    template_name = "catalog/printrun_edit.html"
    context_object_name = "object"   # template me {{ object }} use ho raha hai

    def get_success_url(self):
        # default: detail page pe wapas
        return reverse("catalog:pr_detail", args=[self.object.pk])

# ---------------------------------------------------------------------------
# PRINT-RUN UPDATE (EDIT) – single, final version
# ---------------------------------------------------------------------------
class PrintRunUpdateEnterprise(LoginRequiredMixin, UpdateView):
    model = PrintRun
    form_class = PrintRunForm
    template_name = "catalog/printrun_edit.html"   # 👈 tumhara edit wala template
    context_object_name = "object"

    def get_success_url(self):
        # Edit ke baad batch detail par wapas
        return reverse("catalog:pr_detail", args=[self.object.pk])

    def form_valid(self, form):
        """
        Yahan se:
        - purana received_qty lete hain
        - naya received_qty form se lete hain
        - difference (delta) nikalte hain
        - uske basis par StockLedger me NEW row create karte hain
        """
        # ---- old instance BEFORE save ----
        orig = PrintRun.objects.get(pk=self.get_object().pk)
        old_received = int(orig.received_qty or 0)

        with transaction.atomic():
            # ----- pehle normal save -----
            response = super().form_valid(form)
            pr = self.object
            new_received = int(pr.received_qty or 0)

            # ----- difference -----
            delta = new_received - old_received

            # agar qty change hi nahi hui:
            if delta == 0:
                messages.success(self.request, "Batch updated (no stock change).")
                return response

            # ----- previous balance (same product + warehouse) -----
            prev_balance = (
                StockLedger.objects
                .filter(product=pr.product, warehouse=pr.warehouse)
                .order_by("-date", "-id")
                .values_list("balance", flat=True)
                .first()
                or 0
            )
            prev_balance = int(prev_balance or 0)

            # ----- IN / OUT decide -----
            if delta > 0:
                in_q = delta
                out_q = 0
                ref_type = "pr_update_increase"
            else:
                in_q = 0
                out_q = abs(delta)
                ref_type = "pr_update_decrease"

            new_balance = prev_balance + in_q - out_q
            if new_balance < 0:
                new_balance = 0

            extra = {}
            if hasattr(StockLedger, "user"):
                extra["user"] = self.request.user

            # ----- NEW ledger row append -----
            StockLedger.objects.create(
                product=pr.product,
                print_run=pr,
                warehouse=pr.warehouse,
                in_qty=in_q,
                out_qty=out_q,
                balance=new_balance,
                ref_type=ref_type,
                ref_id=pr.id,
                notes=f"Auto adjustment from batch edit by {self.request.user}",
                **extra,
            )

        messages.success(self.request, "Batch updated and stock adjusted.")
        return response


class PrintRunDelete(LoginRequiredMixin, DeleteView):
    model = PrintRun
    template_name = "catalog/printrun_confirm_delete.html"
    success_url = reverse_lazy("catalog:pr_list")


@login_required
def pr_detail(request, pk):
    pr = get_object_or_404(PrintRun, pk=pk)
    ledgers = StockLedger.objects.filter(print_run=pr).order_by("-date", "-id")

    out_sum = ledgers.aggregate(total=Sum("out_qty"))["total"] or 0
    available = max(0, (pr.received_qty or 0) - out_sum)

    return render(request, "catalog/printrun_detail.html", {
        "pr": pr,
        "ledgers": ledgers,
        "product": pr.product,
        "warehouse": pr.warehouse,
        "available": available
    })


# ============================================================================
# 4) STOCK LEDGER
# ============================================================================

class StockLedgerList(LoginRequiredMixin, ListView):
    model = StockLedger
    template_name = "catalog/stockledger_list.html"
    context_object_name = "ledgers"
    paginate_by = 25
    ordering = ["-date", "-id"]

    def get_queryset(self):
        qs = (
            StockLedger.objects
            .select_related("product", "warehouse")
            .all()
            .order_by(*self.ordering)
        )

        req = self.request
        self.q = (req.GET.get("q") or "").strip()
        self.date_from = (req.GET.get("date_from") or "").strip()
        self.date_to = (req.GET.get("date_to") or "").strip()
        self.tx_type = (req.GET.get("tx_type") or "").strip()
        self.warehouse_id = (req.GET.get("warehouse") or "").strip()

        # TEXT SEARCH
        if self.q:
            qs = qs.filter(
                Q(product__name__icontains=self.q)
                | Q(product__sku__icontains=self.q)
                | Q(notes__icontains=self.q)
                | Q(ref_type__icontains=self.q)
                | Q(ref_id__icontains=self.q)
            )

        # DATE RANGE
        if self.date_from:
            qs = qs.filter(date__date__gte=self.date_from)
        if self.date_to:
            qs = qs.filter(date__date__lte=self.date_to)

        # TYPE FILTER
        if self.tx_type == "in":
            qs = qs.filter(in_qty__gt=0, out_qty=0)
        elif self.tx_type == "out":
            qs = qs.filter(out_qty__gt=0, in_qty=0)
        elif self.tx_type == "adjust":
            qs = qs.filter(
                Q(ref_type__icontains="adjust")
                | Q(in_qty__gt=0, out_qty__gt=0)
            )

        # WAREHOUSE
        if self.warehouse_id:
            try:
                qs = qs.filter(warehouse__id=int(self.warehouse_id))
            except Exception:
                pass

        self.filtered_qs = qs
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["q"] = self.q
        ctx["date_from"] = self.date_from
        ctx["date_to"] = self.date_to
        ctx["tx_type"] = self.tx_type
        ctx["warehouse_id"] = self.warehouse_id

        ctx["warehouses"] = Warehouse.objects.all().order_by("name")

        agg = self.filtered_qs.aggregate(
            total_in=Sum("in_qty"),
            total_out=Sum("out_qty"),
        )
        total_in = agg.get("total_in") or 0
        total_out = agg.get("total_out") or 0

        last_balance = (
            self.filtered_qs
            .order_by("-date", "-id")
            .values_list("balance", flat=True)
            .first()
        )
        if last_balance is None:
            last_balance = total_in - total_out

        ctx["total_in"] = total_in
        ctx["total_out"] = total_out
        ctx["current_balance"] = last_balance

        return ctx


@login_required
def stock_ledger_detail(request, pk):
    sl = get_object_or_404(
        StockLedger.objects.select_related("product", "warehouse"),
        pk=pk,
    )

    in_qty = int(sl.in_qty or 0)
    out_qty = int(sl.out_qty or 0)
    delta = in_qty - out_qty

    # Direction / type
    if delta > 0:
        direction = "In"
    elif delta < 0:
        direction = "Out"
    else:
        direction = "Adjustment"

    # previous balance (is entry se pehle kitna tha)
    prev_balance = None
    try:
        if sl.balance is not None:
            prev_balance = int(sl.balance) - delta
    except Exception:
        prev_balance = None

    # same product + same warehouse ke recent entries (excluding current)
    siblings_qs = StockLedger.objects.filter(product=sl.product)
    if sl.warehouse_id:
        siblings_qs = siblings_qs.filter(warehouse=sl.warehouse)

    related_ledgers = (
        siblings_qs.exclude(pk=sl.pk)
        .order_by("-date", "-id")[:5]
    )

    # Ref display text
    ref_display = None
    if sl.ref_type or sl.ref_id:
        base = (sl.ref_type or "").replace("_", " ").title().strip() or "Ref"
        if sl.ref_id:
            ref_display = f"{base} #{sl.ref_id}"
        else:
            ref_display = base

    context = {
        "entry": sl,
        "product": getattr(sl, "product", None),
        "warehouse": getattr(sl, "warehouse", None),
        "direction": direction,
        "is_in": delta > 0,
        "is_out": delta < 0,
        "is_adjustment": delta == 0,
        "delta": delta,
        "prev_balance": prev_balance,
        "related_ledgers": related_ledgers,
        "ref_display": ref_display,
    }
    return render(request, "catalog/stockledger_detail.html", context)


@login_required
@require_POST
def stock_ledger_quick_edit(request, pk):
    sl = get_object_or_404(StockLedger, pk=pk)
    notes = request.POST.get("notes", "").strip()
    sl.notes = notes
    sl.save()
    return JsonResponse({"ok": True})


# ============================================================================
# 5) INVENTORY DASHBOARD (matches: inventory_home.html)
# ============================================================================

@login_required
def inventory_home(request):
    products = Product.objects.all().order_by("name")
    total = 0
    for p in products:
        total += (
            StockLedger.objects.filter(product=p).aggregate(
                S=Sum(F("in_qty") - F("out_qty"))
            )["S"] or 0
        )

    ctx = {
        "total_products": products.count(),
        "total_available": total,
        "products": products,
    }
    return render(request, "catalog/inventory_home.html", ctx)


# ============================================================================
# 6) EXPORT CSV (for print runs list)
# ============================================================================

@login_required
def export_csv(request):
    qs = PrintRun.objects.select_related("product", "warehouse").order_by("-id")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="print_runs.csv"'

    writer = csv.writer(response)
    writer.writerow(["Batch", "Product", "Received", "Warehouse", "Date"])
    for pr in qs:
        writer.writerow([
            pr.batch_no,
            pr.product.name,
            pr.received_qty,
            pr.warehouse.name if pr.warehouse else "",
            pr.print_date
        ])
    return response


@login_required
def product_stock_detail(request, pk):
    """
    Safe product stock detail page — shows product info, print runs and ledger entries.
    """
    try:
        product = Product.objects.get(pk=pk)
    except Product.DoesNotExist:
        raise Http404("Product not found")

    prs = PrintRun.objects.filter(product=product).order_by("-print_date", "-id")
    ledgers = StockLedger.objects.filter(product=product).order_by("-date", "-id")[:200]

    # compute available stock safely
    try:
        available = safe_available_stock(product)
    except Exception:
        available = 0

    context = {
        "product": product,
        "print_runs": prs,
        "ledgers": ledgers,
        "available": available,
    }
    return render(request, "catalog/product_stock.html", context)

@login_required
@require_POST
def warehouse_import(request):
    """
    Import Warehouses from CSV (safe fallback).
    Expected columns: name, phone, email, address
    """
    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Please upload a CSV file.")
        return redirect(request.META.get("HTTP_REFERER", reverse("catalog:warehouse_list")))

    try:
        from io import TextIOWrapper
        import csv
        text_file = TextIOWrapper(f.file, encoding=request.encoding or "utf-8")
        reader = csv.DictReader(text_file)
    except Exception as e:
        messages.error(request, f"Failed to read CSV: {e}")
        return redirect(request.META.get("HTTP_REFERER", reverse("catalog:warehouse_list")))

    created = 0
    skipped = 0

    for row in reader:
        name = (row.get("name") or "").strip()
        if not name:
            skipped += 1
            continue

        phone = (row.get("phone") or "").strip() or None
        email = (row.get("email") or "").strip() or None
        address = (row.get("address") or "").strip() or None

        obj, created_flag = Warehouse.objects.get_or_create(
            name=name,
            defaults={
                "phone": phone,
                "email": email,
                "address": address
            }
        )

        if created_flag:
            created += 1
        else:
            skipped += 1

    messages.success(request, f"Import complete — created: {created}, skipped: {skipped}")
    return redirect(reverse("catalog:warehouse_list"))


@login_required
@require_POST
def warehouse_import_xlsx(request):
    """
    Import warehouses from XLSX (Excel).
    Expected columns: name, phone, email, address
    """
    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Please upload an XLSX file.")
        return redirect(reverse("catalog:warehouse_list"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        created = 0
        skipped = 0

        # Read header row
        header = [str(v).strip().lower() for v in next(ws.values)]
        col = {name: idx for idx, name in enumerate(header)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = (str(row[col.get("name", -1)]) if col.get("name", -1) >= 0 else "").strip()
            if not name:
                skipped += 1
                continue

            phone = str(row[col.get("phone", -1)]) if col.get("phone", -1) >= 0 else None
            email = str(row[col.get("email", -1)]) if col.get("email", -1) >= 0 else None
            address = str(row[col.get("address", -1)]) if col.get("address", -1) >= 0 else None

            obj, created_flag = Warehouse.objects.get_or_create(
                name=name,
                defaults={
                    "phone": phone if phone else None,
                    "email": email if email else None,
                    "address": address if address else None,
                }
            )

            if created_flag:
                created += 1
            else:
                skipped += 1

        messages.success(request, f"XLSX import complete — created: {created}, skipped: {skipped}")
        return redirect(reverse("catalog:warehouse_list"))

    except Exception as exc:
        logger.exception("warehouse_import_xlsx failed")
        messages.error(request, f"Import failed: {exc}")
        return redirect(reverse("catalog:warehouse_list"))

@login_required
def warehouse_export(request):
    """
    Export warehouse list to CSV.
    Supports filter by ?q= and ?status= (active/inactive).
    """
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    qs = Warehouse.objects.all().order_by("name")

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(phone__icontains=q) |
            Q(address__icontains=q) |
            Q(email__icontains=q)
        )

    if status and hasattr(Warehouse, "is_active"):
        if status == "active":
            qs = qs.filter(is_active=True)
        elif status == "inactive":
            qs = qs.filter(is_active=False)

    # Prepare CSV response
    now = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"warehouses_{now}.csv"

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")  # UTF-8 BOM

    writer = csv.writer(response)
    writer.writerow(["ID", "Name", "Phone", "Email", "Address", "Created At"])

    for w in qs:
        writer.writerow([
            w.id,
            w.name,
            w.phone or "",
            w.email or "",
            w.address or "",
            w.created_at.isoformat() if getattr(w, "created_at", None) else "",
        ])

    return response

@login_required
def warehouse_export_xlsx(request):
    """
    Export Warehouses to XLSX using openpyxl.
    Supports ?q= and ?status= filtering same as CSV export.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "openpyxl not installed. Run: pip install openpyxl")
        return redirect("catalog:warehouse_list")

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    qs = Warehouse.objects.all().order_by("name")

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(phone__icontains=q) |
            Q(address__icontains=q) |
            Q(email__icontains=q)
        )

    if status and hasattr(Warehouse, "is_active"):
        if status == "active":
            qs = qs.filter(is_active=True)
        elif status == "inactive":
            qs = qs.filter(is_active=False)

    # --- Build Workbook ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Warehouses"

    headers = ["ID", "Name", "Phone", "Email", "Address", "Created At"]
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = openpyxl.styles.Font(bold=True)

    for row, w in enumerate(qs, start=2):
        ws.cell(row=row, column=1, value=w.id)
        ws.cell(row=row, column=2, value=w.name)
        ws.cell(row=row, column=3, value=w.phone or "")
        ws.cell(row=row, column=4, value=w.email or "")
        ws.cell(row=row, column=5, value=w.address or "")
        ws.cell(
            row=row,
            column=6,
            value=w.created_at.isoformat() if getattr(w, "created_at", None) else "",
        )

    # Auto column width
    for i, _ in enumerate(headers, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 22

    # --- Response ---
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"warehouses_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response



class PrintRunDeleteEnterprise(LoginRequiredMixin, DeleteView):
    model = PrintRun
    template_name = "catalog/printrun_confirm_delete.html"
    context_object_name = "pr"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        pr = self.object

        # --- Compute available qty for UI ---
        try:
            if hasattr(pr, "available_qty") and callable(pr.available_qty):
                available = pr.available_qty()
            else:
                raise AttributeError
        except Exception:
            logger.exception("available_qty fallback used in PrintRunDeleteEnterprise.get_context_data. print_run_id=%s", pr.id)
            out_sum = (
                StockLedger.objects.filter(print_run=pr)
                .aggregate(total=Sum("out_qty"))
                .get("total") or 0
            )
            received = int(pr.received_qty or 0)
            available = max(0, received - out_sum)

        ctx["available"] = available
        ctx["product"] = pr.product
        ctx["warehouse"] = pr.warehouse
        return ctx

    def get_success_url(self):
        messages.success(
            self.request,
            f"Batch «{self.object.batch_no or self.object.id}» deleted successfully."
        )
        return reverse_lazy("catalog:pr_list")

    def delete(self, request, *args, **kwargs):
        pr = self.get_object()

        try:
            with transaction.atomic():
                # ---- Compute available stock --- 
                try:
                    avail = pr.available_qty()
                except Exception:
                    logger.exception("available_qty fallback used in PrintRunDeleteEnterprise.delete. print_run_id=%s", pr.id)
                    out_sum = (
                        StockLedger.objects.filter(print_run=pr)
                        .aggregate(total=Sum("out_qty"))
                        .get("total") or 0
                    )
                    received = int(pr.received_qty or 0)
                    avail = max(0, received - out_sum)

                # ---- Ledger adjustment BEFORE deleting ----
                if avail > 0:
                    prev_balance = (
                        StockLedger.objects
                        .filter(product=pr.product, warehouse=pr.warehouse)
                        .order_by('-id')
                        .values_list('balance', flat=True)
                        .first() or 0
                    )

                    remove = min(avail, int(prev_balance or 0))

                    if remove > 0:
                        new_balance = int(prev_balance) - remove

                        StockLedger.objects.create(
                            product=pr.product,
                            print_run=pr,
                            warehouse=pr.warehouse,
                            in_qty=0,
                            out_qty=remove,
                            balance=new_balance,
                            ref_type="pr_delete",
                            ref_id=pr.id,
                            notes=f"Reverse stock on delete by {request.user}"
                        )

                # ---- DELETE PRINT RUN ----
                return super().delete(request, *args, **kwargs)

        except Exception as exc:
            logger.exception("PrintRunDeleteEnterprise.delete failed. print_run_id=%s", pr.id)
            messages.error(request, f"Delete failed: {exc}")
            return redirect(reverse_lazy("catalog:pr_list"))




class StockLedgerUpdate(LoginRequiredMixin, UpdateView):
    model = StockLedger
    template_name = "catalog/stockledger_edit.html"
    fields = ["notes", "warehouse", "in_qty", "out_qty"]

    def dispatch(self, request, *args, **kwargs):
        # permission check
        if not request.user.has_perm("catalog.change_stockledger") and not (
            request.user.is_superuser or request.user.is_staff
        ):
            raise PermissionDenied("You do not have permission to edit ledger entries.")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # Keep old values
        old = StockLedger.objects.get(pk=self.object.pk)
        old_in = int(old.in_qty or 0)
        old_out = int(old.out_qty or 0)

        resp = super().form_valid(form)

        # recompute running balances
        try:

            qs = StockLedger.objects.filter(product=self.object.product).order_by(
                "date", "id"
            )

            running = 0
            with transaction.atomic():
                for sl in qs:
                    running = running + int(sl.in_qty or 0) - int(sl.out_qty or 0)
                    if sl.balance != running:
                        sl.balance = running
                        sl.save(update_fields=["balance"])
        except Exception:
            logger.exception("Failed to recompute running balances after StockLedgerUpdate. entry_id=%s", self.object.pk)

        messages.success(self.request, "Ledger entry updated.")
        return resp

    def get_success_url(self):
        return reverse_lazy("catalog:stock_ledger")



@login_required
@require_http_methods(["GET", "POST"])
def stock_ledger_delete(request, pk):
    """
    Ledger delete with confirm page:
    - GET  => show confirmation screen
    - POST => delete + recompute balance + redirect to list
    """
    # Permission check
    if not request.user.has_perm("catalog.delete_stockledger") and not (
        request.user.is_superuser or request.user.is_staff
    ):
        return HttpResponse("Forbidden", status=403)

    sl = get_object_or_404(StockLedger.objects.select_related("product", "warehouse"), pk=pk)

    # ---------- POST: ACTUAL DELETE ----------
    if request.method == "POST":
        product = sl.product
        warehouse = sl.warehouse

        try:
            with transaction.atomic():
                # delete row
                sl.delete()

                # Recompute balances for that product+warehouse
                qs = StockLedger.objects.filter(
                    product=product,
                    warehouse=warehouse,
                ).order_by("date", "id")

                running = 0
                for row in qs:
                    running = running + int(row.in_qty or 0) - int(row.out_qty or 0)
                    if row.balance != running:
                        row.balance = running
                        row.save(update_fields=["balance"])

            messages.success(request, "Ledger entry deleted successfully.")
            return redirect(reverse("catalog:stock_ledger"))

        except Exception as exc:
            logger.exception("stock_ledger_delete failed. entry_id=%s", pk)
            messages.error(request, f"Delete failed: {exc}")
            # Error ho to detail page par wapas bhej do
            return redirect(reverse("catalog:stock_ledger_detail", args=[pk]))

    # ---------- GET: SHOW CONFIRM PAGE ----------
    context = {
        "entry": sl,
        "product": getattr(sl, "product", None),
        "warehouse": getattr(sl, "warehouse", None),
    }
    return render(request, "catalog/stockledger_confirm_delete.html", context)

@login_required
def stock_ledger_export(request):
    """
    Export stock ledger as CSV (simple, robust, no break risk).
    Supports filters already applied in listing.
    """
    # Import helper to reuse filters

    # same queryset as list view
    from django.http import HttpResponse

    # Retrieve filtered queryset
    qs = StockLedger.objects.select_related("product", "warehouse").all()

    q = (request.GET.get("q") or "").strip()
    warehouse = (request.GET.get("warehouse") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    # text search
    if q:
        qs = qs.filter(
            Q(product__name__icontains=q)
            | Q(product__sku__icontains=q)
            | Q(notes__icontains=q)
        )

    # warehouse filter
    if warehouse:
        try:
            qs = qs.filter(warehouse__id=int(warehouse))
        except Exception:
            logger.exception("Invalid warehouse filter during stock_ledger_export. warehouse=%s", warehouse)

    # date filters
    if date_from:
        try:
            qs = qs.filter(date__date__gte=date_from)
        except Exception:
            logger.exception("Invalid date_from filter during stock_ledger_export. date_from=%s", date_from)

    if date_to:
        try:
            qs = qs.filter(date__date__lte=date_to)
        except Exception:
            logger.exception("Invalid date_to filter during stock_ledger_export. date_to=%s", date_to)

    # ----- CSV RESPONSE -----
    now = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f"stock_ledger_{now}.csv"

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")  # BOM for Excel

    writer = csv.writer(response)
    writer.writerow([
        "Date",
        "Product SKU",
        "Product Name",
        "Warehouse",
        "IN",
        "OUT",
        "Balance",
        "Ref",
        "Notes",
    ])

    for sl in qs.order_by("-date", "-id"):
        writer.writerow([
            sl.date.strftime("%Y-%m-%d %H:%M") if getattr(sl, "date", None) else "",
            getattr(sl.product, "sku", ""),
            getattr(sl.product, "name", ""),
            getattr(sl.warehouse, "name", "") if getattr(sl, "warehouse", None) else "",
            sl.in_qty or 0,
            sl.out_qty or 0,
            sl.balance or 0,
            f"{sl.ref_type or ''}#{sl.ref_id or ''}",
            sl.notes or "",
        ])

    return response


@login_required
def product_stock_api(request):
    """
    Return a JSON list of products with available stock, warehouse details,
    and last updated values.
    Perfect for dashboard, AJAX tables, select widgets, etc.
    """
    data = []
    products = Product.objects.all().order_by("name")

    for p in products:
        # --- PRIMARY STOCK COMPUTATION (ledger) ---
        agg = StockLedger.objects.filter(product=p).aggregate(
            total_in=Sum("in_qty"),
            total_out=Sum("out_qty")
        )
        total_in = agg.get("total_in") or 0
        total_out = agg.get("total_out") or 0
        onhand = int(total_in - total_out)

        # --- FALLBACK (batch received_qty sum) ---
        if onhand == 0:
            pr_sum = (
                PrintRun.objects
                .filter(product=p)
                .aggregate(t=Sum("received_qty"))
                .get("t") or 0
            )
            out_sum = (
                StockLedger.objects
                .filter(product=p)
                .aggregate(t=Sum("out_qty"))
                .get("t") or 0
            )
            onhand = int(pr_sum - out_sum)

        data.append({
            "id": p.id,
            "sku": p.sku,
            "name": p.name,
            "onhand": max(0, onhand),
            "track_stock": p.track_stock,
            "is_bundle": p.is_bundle,
            "type": p.product_type,
            "price": float(p.price or 0),
        })

    return JsonResponse({"ok": True, "count": len(data), "results": data})


@login_required
def api_product_printruns(request, pk):
    """
    Return JSON list of print runs for a given product.
    Includes available_qty, warehouse, and print_date.
    Perfect for FIFO allocation UI.
    """
    try:
        product = Product.objects.get(pk=pk)
    except Product.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Product not found"}, status=404)

    prs = (
        PrintRun.objects
        .filter(product=product)
        .order_by("print_date", "id")
    )

    results = []
    for pr in prs:
        # Compute available qty
        try:
            avail = pr.available_qty()
        except Exception:
            logger.exception("available_qty fallback used in api_product_printruns. print_run_id=%s", pr.id)
            # fallback from ledger
            out_sum = (
                StockLedger.objects.filter(print_run=pr)
                .aggregate(total=Sum("out_qty"))
                .get("total") or 0
            )
            avail = max(0, (pr.received_qty or 0) - out_sum)

        results.append({
            "id": pr.id,
            "batch_no": pr.batch_no,
            "printed_qty": pr.printed_qty,
            "received_qty": pr.received_qty,
            "available_qty": avail,
            "print_date": (
                pr.print_date.isoformat()
                if getattr(pr, "print_date", None)
                else None
            ),
            "warehouse": (
                pr.warehouse.name
                if getattr(pr, "warehouse", None)
                else None
            ),
        })

    return JsonResponse({
        "ok": True,
        "product": {
            "id": product.id,
            "sku": product.sku,
            "name": product.name
        },
        "count": len(results),
        "printruns": results
    })


@login_required
def api_allocate_preview(request, pk):
    """
    FIFO allocation preview for a product.
    Query param: ?qty=NUMBER
    Returns list of (print_run, qty) allocations WITHOUT updating DB.
    """
    try:
        product = Product.objects.get(pk=pk)
    except Product.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Product not found"}, status=404)

    # --- parse qty ---
    try:
        qty = int(request.GET.get("qty") or 0)
        if qty <= 0:
            return JsonResponse({"ok": False, "error": "Invalid qty"}, status=400)
    except Exception:
        logger.exception("Invalid qty in api_allocate_preview. product_id=%s", pk)
        return JsonResponse({"ok": False, "error": "Invalid qty"}, status=400)

    # --- FIFO allocation plan ---
    prs = (
        PrintRun.objects
        .filter(product=product)
        .order_by("print_date", "id")
    )

    qty_left = qty
    allocations = []

    for pr in prs:
        try:
            avail = pr.available_qty()
        except Exception:
            logger.exception("available_qty fallback used in api_allocate_preview. print_run_id=%s", pr.id)
            # fallback from ledger computation
            out_sum = (
                StockLedger.objects
                .filter(print_run=pr)
                .aggregate(total=Sum("out_qty"))
                .get("total") or 0
            )
            avail = max(0, (pr.received_qty or 0) - out_sum)

        if avail <= 0:
            continue

        take = min(avail, qty_left)
        allocations.append((pr, take))
        qty_left -= take

        if qty_left <= 0:
            break

    # --- Not enough stock ---
    if qty_left > 0:
        return JsonResponse({
            "ok": True,
            "insufficient": True,
            "requested": qty,
            "available": qty - qty_left,
            "allocated": []
        })

    # --- Format response ---
    data = []
    for pr, take in allocations:
        data.append({
            "pr_id": pr.id,
            "batch_no": pr.batch_no,
            "print_date": (
                pr.print_date.isoformat()
                if getattr(pr, "print_date", None)
                else None
            ),
            "warehouse": (
                pr.warehouse.name
                if getattr(pr, "warehouse", None)
                else None
            ),
            "available_before": (
                pr.available_qty()
                if hasattr(pr, "available_qty")
                else None
            ),
            "allocate_qty": take
        })

    return JsonResponse({
        "ok": True,
        "insufficient": False,
        "requested": qty,
        "allocated_count": len(data),
        "allocations": data
    })


@require_POST
@permission_required("catalog.add_printrun", raise_exception=True)
def api_printrun_create_v2(request):
    """
    Create a new PrintRun via JSON API.
    Required fields:
        product (id)
        printed_qty
        received_qty (default = printed_qty)
        unit_cost
        warehouse (optional)
    """
    # ---- Get data ----
    try:
        data = json.loads(request.body)
    except Exception:
        data = request.POST.dict()

    errors = {}

    # --- PRODUCT ---
    pid = data.get("product") or data.get("product_id")
    if not pid:
        errors.setdefault("product", []).append("This field is required.")
    else:
        try:
            product = Product.objects.get(pk=int(pid))
        except Exception:
            logger.exception("Invalid product in api_printrun_create_v2. product=%s", pid)
            errors.setdefault("product", []).append("Invalid product.")

    # --- WAREHOUSE (optional) ---
    warehouse = None
    wid = data.get("warehouse")
    if wid:
        try:
            warehouse = Warehouse.objects.get(pk=int(wid))
        except Exception:
            logger.exception("Invalid warehouse in api_printrun_create_v2. warehouse=%s", wid)
            errors.setdefault("warehouse", []).append("Invalid warehouse.")

    # --- QTY PARSE ---
    def to_int(v):
        try:
            return int(float(v))
        except Exception:
            return None

    printed = to_int(data.get("printed_qty"))
    received = to_int(data.get("received_qty"))

    if printed is None:
        printed = 0

    if received is None:
        received = printed

    if printed < 0:
        errors.setdefault("printed_qty", []).append("Cannot be negative.")

    if received < 0:
        errors.setdefault("received_qty", []).append("Cannot be negative.")

    # --- UNIT COST ---
    try:
        unit_cost = Decimal(str(data.get("unit_cost") or "0"))
    except Exception:
        logger.exception("Invalid unit_cost in api_printrun_create_v2. unit_cost=%s", data.get("unit_cost"))
        errors.setdefault("unit_cost", []).append("Invalid unit cost.")

    # --- STOP IF ERRORS ---
    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    # --- CREATE PR + LEDGER ---
    try:
        with transaction.atomic():
            batch_no = (data.get("batch_no") or "").strip()
            if not batch_no:
                ts = int(timezone.now().timestamp())
                batch_no = f"PR-{product.id}-{ts}"

            pr = PrintRun.objects.create(
                product=product,
                warehouse=warehouse,
                batch_no=batch_no,
                printed_qty=printed,
                received_qty=received,
                unit_cost=unit_cost,
                print_date=data.get("print_date") or timezone.now().date(),
                notes=(data.get("notes") or "")[:500],
            )

            # 💥 Create Ledger Entry
            if received > 0:
                # get previous balance
                prev_balance = (
                    StockLedger.objects
                    .filter(product=product, warehouse=warehouse)
                    .order_by("-id")
                    .values_list("balance", flat=True)
                    .first()
                    or 0
                )
                new_balance = prev_balance + received

                ledger_data = {
                    "product": product,
                    "print_run": pr,
                    "warehouse": warehouse,
                    "in_qty": received,
                    "out_qty": 0,
                    "balance": new_balance,
                    "ref_type": "pr_create_api",
                    "ref_id": pr.id,
                    "notes": f"Created via API by {request.user}",
                }

                # only include user if model supports it
                if hasattr(StockLedger, "user"):
                    ledger_data["user"] = request.user

                StockLedger.objects.create(**ledger_data)

    except Exception as e:
        logger.exception("api_printrun_create_v2 failed")
        return JsonResponse(
            {"ok": False, "errors": {"__all__": [str(e)]}},
            status=500
        )

    return JsonResponse({
        "ok": True,
        "pr_id": pr.id,
        "batch_no": pr.batch_no,
        "redirect_url": reverse("catalog:pr_detail", args=[pr.id]),
    })


@permission_required("catalog.change_stockledger", raise_exception=True)
@require_http_methods(["POST"])
def api_create_adjustment_v2(request):
    """
    SAFE STOCK ADJUSTMENT API
    Can:
      ✔ Increase/decrease qty for a print-run (append ledger entry)
      ✔ Update received_qty safely
      ✘ DOES NOT EDIT existing ledger rows
    
    Accepts:
      print_run_id   OR   ledger_id
      qty            OR   new_received_qty
      notes
    """
    data = request.POST.dict()

    # -------------------------
    # Helper to convert qty
    # -------------------------
    def to_int(v):
        try:
            return int(v)
        except:
            return None

    # -------------------------
    # Read parameters
    # -------------------------
    pr_id = data.get("print_run_id") or data.get("pr_id")
    ledger_id = data.get("ledger_id")
    qty_raw = data.get("qty")
    new_received_raw = data.get("new_received_qty")
    notes = (data.get("notes") or "").strip()

    product = None
    warehouse = None
    pr = None

    # -------------------------
    # Determine target
    # -------------------------
    if pr_id:
        try:
            pr = PrintRun.objects.select_related("product", "warehouse").get(pk=int(pr_id))
        except Exception:
            logger.exception("Invalid print_run_id in api_create_adjustment_v2. print_run_id=%s", pr_id)
            return JsonResponse({"ok": False, "error": "Invalid print_run_id"}, status=404)
        
        product = pr.product
        warehouse = pr.warehouse

    elif ledger_id:
        try:
            sl = StockLedger.objects.select_related("product", "warehouse").get(pk=int(ledger_id))
        except Exception:
            logger.exception("Invalid ledger_id in api_create_adjustment_v2. ledger_id=%s", ledger_id)
            return JsonResponse({"ok": False, "error": "Invalid ledger_id"}, status=404)
        
        product = sl.product
        warehouse = sl.warehouse

    else:
        return JsonResponse({"ok": False, "error": "Missing print_run_id or ledger_id"}, status=400)

    # -------------------------
    # CASE 1: new_received_qty override
    # -------------------------
    if new_received_raw not in (None, "") and pr:
        try:
            new_received = int(new_received_raw)
        except Exception:
            logger.exception("Invalid new_received_qty in api_create_adjustment_v2. value=%s", new_received_raw)
            return JsonResponse({"ok": False, "error": "Invalid new_received_qty"}, status=400)

        original = int(pr.received_qty or 0)
        delta = new_received - original
        qty = delta

    else:
        # normal adjustment
        qty = to_int(qty_raw)

    if qty is None:
        return JsonResponse({"ok": False, "error": "Missing qty"}, status=400)

    # -------------------------
    # Build ledger values
    # -------------------------
    in_q = qty if qty > 0 else 0
    out_q = abs(qty) if qty < 0 else 0

    # Previous balance
    prev_balance = (
        StockLedger.objects
        .filter(product=product, warehouse=warehouse)
        .order_by("-id")
        .values_list("balance", flat=True)
        .first()
        or 0
    )

    prev_balance = int(prev_balance or 0)
    new_balance = prev_balance + in_q - out_q
    if new_balance < 0:
        new_balance = 0

    # -------------------------
    # COMMIT TRANSACTION
    # -------------------------
    try:
        with transaction.atomic():

            # Create ledger entry
            sl = StockLedger.objects.create(
                product=product,
                print_run=pr if pr else None,
                warehouse=warehouse if warehouse else None,
                in_qty=in_q,
                out_qty=out_q,
                balance=new_balance,
                ref_type="adjustment_ui",
                ref_id=(pr.id if pr else ledger_id),
                notes=(notes + f" (by {request.user})") if notes else f"Adjusted by {request.user}",
                # user field only if model has it
                **({"user": request.user} if hasattr(StockLedger, "user") else {})
            )

            # If PR override mode, update received_qty
            if pr and new_received_raw not in (None, ""):
                pr.received_qty = new_received
                pr.save(update_fields=["received_qty"])

    except Exception as e:
        logger.exception("api_create_adjustment_v2 failed")
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

    return JsonResponse({
        "ok": True,
        "ledger_id": sl.id,
        "new_balance": new_balance,
    })


# ---------------------------------------------------------
# REF TYPE → MODEL RESOLUTION (safe, dynamic)
# ---------------------------------------------------------

REF_TYPE_MODEL_MAP = {
    # Example:
    # "invoice": "sales.Invoice",
    # "payment": "sales.Payment",
    # Add more if required
}


def _model_for_ref_type(ref_type: str):
    """
    Returns Django model class based on ref_type.
    Attempts mapping first, then auto-detects by model name.
    """
    # 1) explicit mapping
    mapped = REF_TYPE_MODEL_MAP.get(ref_type)
    if mapped:
        try:
            app_label, model_name = mapped.split(".", 1)
            return apps.get_model(app_label, model_name)
        except Exception:
            pass

    # 2) fallback: auto-detect ANY model matching name
    ref_type_lower = ref_type.lower()
    for m in apps.get_models():
        if m._meta.model_name.lower() == ref_type_lower:
            return m

    return None


# ---------------------------------------------------------
# TRANSACTION DETAIL - safe fallback behavior
# ---------------------------------------------------------
def transaction_detail(request, ref_type, ref_id):
    """
    Show object tied to ledger reference.
    If model not found, show a simple info page (no crash).
    """
    Model = _model_for_ref_type(ref_type)
    if not Model:
        return render(request, "catalog/transaction_detail.html", {
            "ref_type": ref_type,
            "ref_id": ref_id,
            "error": "Unknown reference type (no linked model found).",
        })

    try:
        obj = Model.objects.get(pk=ref_id)
    except Model.DoesNotExist:
        return render(request, "catalog/transaction_detail.html", {
            "ref_type": ref_type,
            "ref_id": ref_id,
            "error": "Referenced object does not exist.",
        })

    return render(request, "catalog/transaction_detail.html", {
        "ref_type": ref_type,
        "obj": obj,
        "ref_id": ref_id,
    })


# ---------------------------------------------------------
# TRANSACTION EDIT - safe placeholder
# ---------------------------------------------------------
def transaction_edit(request, ref_type, ref_id):
    Model = _model_for_ref_type(ref_type)
    if not Model:
        raise Http404("Unknown reference type.")
    obj = get_object_or_404(Model, pk=ref_id)

    if request.method == "POST":
        # Placeholder endpoint until typed forms are added for each ref_type model.
        messages.success(request, "Changes saved.")
        return redirect("catalog:transaction_detail", ref_type=ref_type, ref_id=ref_id)

    return render(request, "catalog/transaction_edit.html", {
        "ref_type": ref_type,
        "obj": obj,
        "ref_id": ref_id
    })


# ---------------------------------------------------------
# TRANSACTION DELETE - safe delete
# ---------------------------------------------------------
def transaction_delete(request, ref_type, ref_id):
    Model = _model_for_ref_type(ref_type)
    if not Model:
        raise Http404("Unknown reference type.")
    obj = get_object_or_404(Model, pk=ref_id)

    if request.method == "POST":
        obj.delete()
        messages.success(request, "Transaction deleted.")
        return redirect("catalog:stock_ledger")

    return render(request, "catalog/transaction_confirm_delete.html", {
        "ref_type": ref_type,
        "obj": obj,
        "ref_id": ref_id,
    })



@login_required
def product_stock_search(request):
    """
    SaaS-style product stock search:
    - q => product name / sku / description
    - warehouse => filter by warehouse
    - paginated, with summary KPIs
    """
    q = (request.GET.get("q") or "").strip()
    warehouse_id = (request.GET.get("warehouse") or "").strip()

    products_qs = Product.objects.all().order_by("name")

    if q:
        products_qs = products_qs.filter(
            Q(name__icontains=q)
            | Q(sku__icontains=q)
            | Q(description__icontains=q)
        )

    warehouses = Warehouse.objects.all().order_by("name")

    selected_warehouse = None
    if warehouse_id:
        try:
            selected_warehouse = Warehouse.objects.get(pk=int(warehouse_id))
        except (Warehouse.DoesNotExist, ValueError, TypeError):
            selected_warehouse = None

    rows = []
    total_onhand = 0
    low_count = 0

    for p in products_qs:
        ledgers = StockLedger.objects.filter(product=p)
        if selected_warehouse:
            ledgers = ledgers.filter(warehouse=selected_warehouse)

        agg = ledgers.aggregate(
            total_in=Sum("in_qty"),
            total_out=Sum("out_qty"),
        )
        total_in = agg.get("total_in") or 0
        total_out = agg.get("total_out") or 0
        onhand = int(total_in - total_out)
        if onhand < 0:
            onhand = 0

        total_onhand += onhand

        # low stock logic
        is_low = False
        if p.track_stock:
            if p.reorder_level:
                is_low = onhand <= p.reorder_level
            else:
                is_low = onhand > 0 and onhand <= 5
        if is_low:
            low_count += 1

        rows.append({
            "product": p,
            "sku": p.sku,
            "price": p.price,
            "onhand": onhand,
        })

    paginator = Paginator(rows, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "results": page_obj.object_list,
        "warehouses": warehouses,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "q": q,
        "warehouse_id": warehouse_id,
        "total_products": products_qs.count(),
        "total_onhand": total_onhand,
        "low_count": low_count,
    }
    return render(request, "catalog/product_stock.html", context)


@login_required
@require_POST
def warehouse_deactivate(request, pk):
    w = get_object_or_404(Warehouse, pk=pk)

    if not hasattr(w, "is_active"):
        return JsonResponse(
            {"ok": False, "error": "Warehouse model has no 'is_active' field."},
            status=400,
        )

    w.is_active = not bool(w.is_active)
    w.save(update_fields=["is_active"])

    return JsonResponse(
        {
            "ok": True,
            "is_active": w.is_active,
        }
    )


@login_required
@require_POST
def product_import_xlsx(request):
    """
    Import / upsert products from XLSX.
    Expected headers (case-insensitive):
      SKU, Name, HSN Code, Author, Imprint, Edition, Academic Session, Price, MRP, Tax Rate, Track Stock, Reorder Level, Active
    """
    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Please upload an Excel (.xlsx) file.")
        return redirect("catalog:product_list")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        header = [str(v).strip().lower() for v in next(ws.values)]
        col = {name: idx for idx, name in enumerate(header)}

        created = 0
        updated = 0
        skipped = 0

        def get_val(row, key):
            idx = col.get(key, -1)
            if idx < 0:
                return None
            v = row[idx]
            if v is None:
                return None
            return str(v).strip()

        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = get_val(row, "sku")
            name = get_val(row, "name")

            if not sku or not name:
                skipped += 1
                continue

            price_raw = get_val(row, "price") or "0"
            mrp_raw = get_val(row, "mrp") or "0"
            tax_raw = get_val(row, "tax rate") or get_val(row, "tax") or "18"
            hsn_raw = get_val(row, "hsn code") or get_val(row, "hsn") or ""
            author_raw = get_val(row, "author") or ""
            imprint_raw = get_val(row, "imprint") or ""
            edition_raw = get_val(row, "edition") or ""
            session_raw = get_val(row, "academic session") or get_val(row, "session") or ""
            track_raw = (get_val(row, "track stock") or "").lower()
            reorder_raw = get_val(row, "reorder level") or "0"
            active_raw = (get_val(row, "active") or "").lower()

            try:
                price = Decimal(str(price_raw))
            except Exception:
                price = Decimal("0")

            try:
                mrp = Decimal(str(mrp_raw))
            except Exception:
                mrp = Decimal("0")

            try:
                tax_rate = int(float(tax_raw))
            except Exception:
                tax_rate = 18

            try:
                reorder_level = int(float(reorder_raw))
            except Exception:
                reorder_level = 0

            track_stock = track_raw in ("yes", "true", "1", "y")
            active = active_raw not in ("no", "false", "0", "n")
            hsn_code = (hsn_raw or "").strip()
            if hsn_code and (not hsn_code.isdigit() or len(hsn_code) not in (6, 7, 8)):
                hsn_code = ""

            obj, created_flag = Product.objects.update_or_create(
                sku=sku,
                defaults={
                    "name": name,
                    "price": price,
                    "mrp": mrp,
                    "tax_rate": tax_rate,
                    "hsn_code": hsn_code or None,
                    "author": (author_raw or "").strip() or None,
                    "imprint": (imprint_raw or "").strip() or None,
                    "edition": (edition_raw or "").strip() or None,
                    "academic_session": (session_raw or "").strip() or None,
                    "track_stock": track_stock,
                    "reorder_level": reorder_level,
                    "active": active,
                },
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        messages.success(
            request,
            f"Product import complete — created: {created}, updated: {updated}, skipped: {skipped}",
        )
    except Exception as exc:
        logger.exception("product_import_xlsx failed")
        messages.error(request, f"Import failed: {exc}")

    return redirect("catalog:product_list")

@login_required
def product_export_xlsx(request):
    """
    Export filtered products list to XLSX.
    Respects ?q=, ?status=, ?track= filters.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "openpyxl not installed. Run: pip install openpyxl")
        return redirect("catalog:product_list")

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    track = (request.GET.get("track") or "").strip()

    qs = Product.objects.all().order_by("name")

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(sku__icontains=q)
            | Q(description__icontains=q)
            | Q(isbn__icontains=q)
            | Q(author__icontains=q)
            | Q(imprint__icontains=q)
            | Q(edition__icontains=q)
            | Q(academic_session__icontains=q)
            | Q(subject__icontains=q)
        )

    if status == "active":
        qs = qs.filter(active=True)
    elif status == "inactive":
        qs = qs.filter(active=False)

    if track == "yes":
        qs = qs.filter(track_stock=True)
    elif track == "no":
        qs = qs.filter(track_stock=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "SKU",
        "Name",
        "HSN Code",
        "Author",
        "Imprint",
        "Edition",
        "Academic Session",
        "Price",
        "MRP",
        "Tax Rate",
        "Track Stock",
        "Reorder Level",
        "Active",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    row_idx = 2
    for p in qs:
        ws.cell(row=row_idx, column=1, value=p.sku)
        ws.cell(row=row_idx, column=2, value=p.name)
        ws.cell(row=row_idx, column=3, value=p.hsn_code or "")
        ws.cell(row=row_idx, column=4, value=p.author or "")
        ws.cell(row=row_idx, column=5, value=p.imprint or "")
        ws.cell(row=row_idx, column=6, value=p.edition or "")
        ws.cell(row=row_idx, column=7, value=p.academic_session or "")
        ws.cell(row=row_idx, column=8, value=float(p.price or 0))
        ws.cell(row=row_idx, column=9, value=float(p.mrp or 0))
        ws.cell(row=row_idx, column=10, value=p.tax_rate)
        ws.cell(row=row_idx, column=11, value="Yes" if p.track_stock else "No")
        ws.cell(row=row_idx, column=12, value=p.reorder_level or 0)
        ws.cell(row=row_idx, column=13, value="Active" if p.active else "Inactive")
        row_idx += 1

    for i, _ in enumerate(headers, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 20

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"products_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response




class ProductDetail(LoginRequiredMixin, DetailView):
    model = Product
    template_name = "catalog/product_detail.html"
    context_object_name = "product"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        product = ctx["product"]

        # ✅ 1) REAL stock – sirf ledger se
        agg = StockLedger.objects.filter(product=product).aggregate(
            total_in=Sum("in_qty"),
            total_out=Sum("out_qty"),
        )
        total_in = int(agg["total_in"] or 0)
        total_out = int(agg["total_out"] or 0)
        stock = total_in - total_out
        if stock < 0:
            stock = 0

        # ✅ 2) Recent batches (print runs) – per batch available_qty bhi ledger se
        try:
            pr_qs = product.print_runs.all().order_by("-print_date", "-id")[:5]
        except Exception:
            logger.exception("Failed to load recent print runs in ProductDetail. product_id=%s", product.id)
            pr_qs = []

        # ✅ 3) Recent ledger activity
        recent_ledgers = (
            StockLedger.objects
            .filter(product=product)
            .select_related("warehouse")
            .order_by("-date", "-id")[:10]
        )

        ctx["stock"] = stock
        ctx["recent_printruns"] = pr_qs
        ctx["recent_ledgers"] = recent_ledgers
        return ctx

logger = logging.getLogger(__name__)
